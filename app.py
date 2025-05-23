import streamlit as st
import pandas as pd
import os
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import warnings
from datetime import datetime, date, time, timedelta
import io
import tempfile
import shutil
import re 
import msoffcrypto
import zipfile
 
from supabase import create_client
from dotenv import load_dotenv
load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

warnings.filterwarnings('ignore', category=UserWarning, 
                        message="Cell .* is marked as a date but the serial value .* is outside the limits for dates.*")
class BaseProcessor:
    def __init__(self):
        self.temp_dir = tempfile.mkdtemp()
        
    def __del__(self):
        try:
            if os.path.exists(self.temp_dir):
                shutil.rmtree(self.temp_dir)
        except:
            pass
          
    def process_mobile_number(self, mobile_num):
        if pd.isna(mobile_num) or mobile_num is None or str(mobile_num).strip() == "":
            return ""

        mobile_num = str(mobile_num).strip()
        mobile_num = re.sub(r'\D', '', mobile_num)

        if mobile_num.startswith('639') and len(mobile_num) == 12:
            result = '09' + mobile_num[3:]
            return result

        if mobile_num.startswith('9') and len(mobile_num) == 10:
            result = '0' + mobile_num
            return result

        if mobile_num.startswith('09') and len(mobile_num) == 11:
            return mobile_num

        if mobile_num.startswith('+639') and len(mobile_num) == 13:
            result = '09' + mobile_num[4:]
            return result

        return mobile_num

    def format_date(self, date_value):
        if pd.isna(date_value) or date_value is None:
            return ""
            
        if isinstance(date_value, (datetime, date)):
            return date_value.strftime("%m/%d/%Y")
        
        try:
            date_obj = pd.to_datetime(date_value)
            return date_obj.strftime("%m/%d/%Y")
        except:
            return str(date_value)
    def clean_data(self, df, remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        if not isinstance(df, pd.DataFrame):
            raise ValueError(f"Expected a pandas DataFrame, but got {type(df)}: {df}")
        
        cleaned_df = df.copy()
        
        if remove_blanks: 
            cleaned_df = cleaned_df.dropna(how='all')
        if remove_duplicates:
            cleaned_df = cleaned_df.drop_duplicates()
        if trim_spaces:
            for col in cleaned_df.select_dtypes(include=['object']).columns:
                cleaned_df[col] = cleaned_df[col].str.strip()
                
        cleaned_df = cleaned_df.replace(r'^\s*$', pd.NA, regex=True)
        return cleaned_df
        
    def clean_only(self, file_content, sheet_name, preview_only=False, 
                   remove_duplicates=False, remove_blanks=False, trim_spaces=False, file_name=None):
        try:
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            sheet_names = xls.sheet_names
            df = pd.read_excel(xls, sheet_name=sheet_names[0])

            sanitized_headers = [re.sub(r'[^A-Za-z0-9_]', '_', str(col)) for col in df.columns]
            df.columns = sanitized_headers

            cleaned_df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)

            if preview_only:
                return cleaned_df

            if file_name:
                base_name = os.path.splitext(os.path.basename(file_name))[0]
                output_filename = f"{base_name}.xlsx"
            else:
                output_filename = "CLEANED_DATA.xlsx"

            output_path = os.path.join(self.temp_dir, output_filename)

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                cleaned_df.to_excel(writer, index=False, sheet_name='Sheet1')
                worksheet = writer.sheets['Sheet1']
                for i, col in enumerate(cleaned_df.columns):
                    try:
                        max_len_in_column = cleaned_df[col].astype(str).map(len).max()
                        max_length = max(max_len_in_column, len(str(col))) + 2
                    except:
                        max_length = 15
                    col_letter = get_column_letter(i + 1)
                    worksheet.column_dimensions[col_letter].width = max_length

            with open(output_path, 'rb') as f:
                output_binary = f.read()

            return cleaned_df, output_binary, output_filename

        except Exception as e:
            st.error(f"Error cleaning file: {str(e)}")
            raise

class BPIProcessor(BaseProcessor):
    
    def setup_directories(self, automation_type):
        """Create necessary directories based on automation type"""
        directories = {
            'updates': ["FOR_UPDATES", "BPI_FOR_UPDATES"],
            'uploads': ["FOR_UPLOADS", "BPI_FOR_UPLOADS"],
            'cured_list': ["CURED_LIST", "BPI_FOR_REMARKS", "BPI_FOR_PAYMENTS", "BPI_FOR_OTHERS"]
        }
        
        dirs_to_create = directories.get(automation_type, [])
        created_dirs = {}
        
        for dir_name in dirs_to_create:
            dir_path = os.path.join(self.temp_dir, dir_name)
            os.makedirs(dir_path, exist_ok=True)
            created_dirs[dir_name] = dir_path
            
        return created_dirs
        
    def process_updates_or_uploads(self, file_content, sheet_name, automation_type, preview_only=False,
                                   remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        try:
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            
            required_columns = [
                'LAN', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 
                'PRINCIPAL', 'LPC', 'ADA SHORTAGE', 'UNIT', 'DPD',
                'EMAIL', 'CONTACT NUMBER 1', 'CONTACT NUMBER 2', 'ENDO DATE'
            ]
            
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if preview_only:
                return df
                
            current_date = datetime.now().strftime('%m%d%Y')
            
            if automation_type == 'updates':
                output_filename = f"BPI AUTO CURING FOR UPDATES {current_date}.xlsx"
                input_filename = f"FOR UPDATE {current_date}.xlsx"
                dirs = self.setup_directories('updates')
                folder_key = 'BPI_FOR_UPDATES'
                input_folder_key = 'FOR_UPDATES'
            else: 
                output_filename = f"BPI AUTO CURING FOR UPLOADS {current_date}.xlsx"
                input_filename = f"FOR UPLOAD (NEW ENDO) {current_date}.xlsx"
                dirs = self.setup_directories('uploads')
                folder_key = 'BPI_FOR_UPLOADS'
                input_folder_key = 'FOR_UPLOADS'
            
            input_path = os.path.join(dirs[input_folder_key], input_filename)
            with open(input_path, 'wb') as f:
                f.write(file_content)
                
            column_map = {
                'EMAIL': 'EMAIL_ALS',
                'CONTACT NUMBER 1': 'MOBILE_NO_ALS',
                'CONTACT NUMBER 2': 'MOBILE_ALFES',
                'ENDO DATE': 'DATE REFERRED'
            }
            
            result_df = pd.DataFrame()
            
            for col in ['LAN', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 
                        'ADA SHORTAGE', 'UNIT', 'DPD']:
                if col in df.columns:
                    result_df[col] = df[col].fillna("")
            
            result_df.insert(1, 'CH CODE', result_df['LAN'])
            
            for orig_col, new_col in column_map.items():
                if orig_col in df.columns:
                    if orig_col == 'CONTACT NUMBER 1' or orig_col == 'CONTACT NUMBER 2':
                        result_df[new_col] = df[orig_col].apply(lambda x: "" if pd.isna(x) else self.process_mobile_number(x))
                    elif orig_col == 'ENDO DATE':
                        result_df[new_col] = df[orig_col].apply(lambda x: self.format_date(x) if pd.notnull(x) else "")
                    else:
                        result_df[new_col] = df[orig_col].fillna("")
                else:
                    result_df[new_col] = ""
            
            result_df['LANDLINE_NO_ALFES'] = ""
            
            numeric_cols = ['PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC', 'ADA SHORTAGE']
            for col in numeric_cols:
                if col in result_df.columns:
                    result_df[col] = pd.to_numeric(result_df[col], errors='coerce').fillna(0).round(2)
                    
            final_columns = [
                'LAN', 'CH CODE', 'NAME', 'CTL4', 'PAST DUE', 'PAYOFF AMOUNT', 'PRINCIPAL', 'LPC',
                'ADA SHORTAGE', 'EMAIL_ALS', 'MOBILE_NO_ALS', 'MOBILE_ALFES', 'LANDLINE_NO_ALFES', 
                'DATE REFERRED', 'UNIT', 'DPD'
            ]
            
            for col in final_columns:
                if col not in result_df.columns:
                    result_df[col] = ""
                    
            result_df = result_df[final_columns]
            
            output_path = os.path.join(dirs[folder_key], output_filename)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                for i, col in enumerate(final_columns):
                    max_length = max(
                        result_df[col].astype(str).map(len).max(),
                        len(col)
                    ) + 2
                    col_letter = chr(65 + i) 
                    worksheet.column_dimensions[col_letter].width = max_length
                    
                    if col in numeric_cols:
                        for row in range(2, len(result_df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            cell.number_format = '0.00'
                    
                    if col == 'DATE REFERRED':
                        for row in range(2, len(result_df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            value = cell.value
                            if value:
                                try: 
                                    cell.value = pd.to_datetime(value).strftime("%m/%d/%Y")
                                    cell.number_format = '@'
                                except:
                                    pass
            
            with open(output_path, 'rb') as f:
                output_binary = f.read()
                
            return result_df, output_binary, output_filename
            
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            raise

    def process_updates(self, file_content, sheet_name=None, preview_only=False,
                        remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        return self.process_updates_or_uploads(file_content, sheet_name, 'updates', preview_only,
                                               remove_duplicates, remove_blanks, trim_spaces)
        
    def process_uploads(self, file_content, sheet_name=None, preview_only=False,
                        remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        return self.process_updates_or_uploads(file_content, sheet_name, 'uploads', preview_only,
                                               remove_duplicates, remove_blanks, trim_spaces)
    
    def process_cured_list(self, file_content, sheet_name=None, preview_only=False,
                           remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_input:
            temp_input.write(file_content)
            temp_input_path = temp_input.name
            
        try:
            xls = pd.ExcelFile(temp_input_path)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if preview_only:
                return df
                
            current_date = datetime.now().strftime('%m%d%Y')
            dirs = self.setup_directories('cured_list')
            
            input_file = os.path.join(dirs["CURED_LIST"], f"CURED LIST {current_date}.xlsx")
            shutil.copy(temp_input_path, input_file)
            
            remarks_filename = f"BPI AUTOCURING REMARKS {current_date}.xlsx"
            others_filename = f"BPI AUTOCURING RESHUFFLE {current_date}.xlsx"
            payments_filename = f"BPI AUTOCURING PAYMENT {current_date}.xlsx"
            
            remarks_path = os.path.join(dirs["BPI_FOR_REMARKS"], remarks_filename)
            others_path = os.path.join(dirs["BPI_FOR_OTHERS"], others_filename)
            payments_path = os.path.join(dirs["BPI_FOR_PAYMENTS"], payments_filename)
            
            try:
                source_wb = openpyxl.load_workbook(input_file)
            except FileNotFoundError:
                print(f"Error: The file '{input_file}' was not found.")
                return
            
            dest_wb = openpyxl.Workbook()
            dest_ws = dest_wb.active
            
            headers = ["LAN", "Action Status", "Remark Date", "PTP Date", "Reason For Default", 
                    "Field Visit Date", "Remark", "Next Call Date", "PTP Amount", "Claim Paid Amount", 
                    "Remark By", "Phone No.", "Relation", "Claim Paid Date"]
            
            for col, header in enumerate(headers, 1):
                dest_ws.cell(row=1, column=col).value = header
            
            ws = source_wb.active
            if ws.max_column < 43:
                raise ValueError("File doesn't have the expected number of columns")
            
            current_row = 2
            total_rows = 0
            last_row = ws.max_row
            
            barcode_lookup = {}
            for row in range(2, last_row + 1):
                barcode = ws.cell(row=row, column=1).value
                phone1 = ws.cell(row=row, column=42).value
                phone2 = ws.cell(row=row, column=43).value
                phone1 = str(phone1).strip() if phone1 else ""
                phone2 = str(phone2).strip() if phone2 else ""
                if phone1:
                    phone1 = self.process_mobile_number(phone1)
                if phone2:
                    phone2 = self.process_mobile_number(phone2)
                if barcode:
                    barcode_lookup[barcode] = {
                        'date': ws.cell(row=row, column=3).value,
                        'amount': ws.cell(row=row, column=4).value,
                        'collector': ws.cell(row=row, column=2).value,
                        'phone1': phone1,
                        'phone2': phone2,
                    }
            
            nego_rows = []
            for row in range(2, last_row + 1):
                if (ws.cell(row=row, column=2).value != "SPMADRID" and 
                    (ws.cell(row=row, column=8).value is None or "PTP" not in str(ws.cell(row=row, column=8).value))):
                    nego_rows.append(row)
            
            if nego_rows:
                visible_count = len(nego_rows)
                
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CALL OUTS_PASTDUE"
                
                current_row += visible_count
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(nego_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 3)
            
            ptp_rows = []
            for row in range(2, last_row + 1):
                if (ws.cell(row=row, column=2).value != "SPMADRID" and 
                    ws.cell(row=row, column=8).value is not None and 
                    "PTP" in str(ws.cell(row=row, column=8).value)):
                    ptp_rows.append(row)
            
            if ptp_rows:
                visible_count = len(ptp_rows)
                
                for i, row_idx in enumerate(ptp_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP FF UP - CLIENT ANSWERED AND WILL SETTLE"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(ptp_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 2)
            
            spmadrid_rows = []
            for row in range(2, last_row + 1):
                if ws.cell(row=row, column=2).value == "SPMADRID":
                    spmadrid_rows.append(row)
            
            if spmadrid_rows:
                visible_count = len(spmadrid_rows)
                
                for i, row_idx in enumerate(spmadrid_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PTP NEW - CURED_GHOST"
                
                current_row += visible_count
                
                for i, row_idx in enumerate(spmadrid_rows):
                    barcode = ws.cell(row=row_idx, column=1).value
                    dest_ws.cell(row=current_row + i, column=1).value = barcode
                    dest_ws.cell(row=current_row + i, column=2).value = "PAYMENT - CURED"
                
                current_row += visible_count
                
                total_rows += (visible_count * 2)
            
            final_row_count = total_rows + 1
            
            for row in range(2, final_row_count + 1):
                barcode = dest_ws.cell(row=row, column=1).value
                action_status = dest_ws.cell(row=row, column=2).value
                
                source_data = barcode_lookup.get(barcode, {})
                source_date = source_data.get('date')
                source_amount = source_data.get('amount')
                source_collector = source_data.get('collector')
                source_phone1 = source_data.get('phone1')
                source_phone2 = source_data.get('phone2')
                
                if source_date:
                    try:
                        if hasattr(source_date, 'strftime'): 
                            base_date = source_date
                        else:
                            try:
                                base_date = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                            except:
                                try:
                                    base_date = datetime.strptime(str(source_date), "%Y-%m-%d")
                                except:
                                    base_date = datetime.now()
                    except:
                        base_date = datetime.now()
                    
                    if "PTP NEW" in action_status:
                        time_to_add = time(14, 40, 0)
                    elif "PTP FF" in action_status:
                        time_to_add = time(14, 50, 0)
                    elif "CURED" in action_status:
                        time_to_add = time(15, 0, 0)
                    else:
                        time_to_add = time(0, 0, 0)
                    
                    if not hasattr(base_date, 'time'):
                        base_date = datetime.combine(base_date, time(0, 0, 0))
                    
                    result_date = datetime.combine(base_date.date(), time_to_add)
                    
                    formatted_date = result_date.strftime("%m/%d/%Y %I:%M:%S %p")
                    dest_ws.cell(row=row, column=3).value = formatted_date
                    
                    formatted_date = result_date.strftime("%m/%d/%Y")
                    dest_ws.cell(row=row, column=4).value = formatted_date
                    
                    dest_ws.cell(row=row, column=3).number_format = '@'
                    dest_ws.cell(row=row, column=4).number_format = '@'
                else:
                    dest_ws.cell(row=row, column=3).value = ""
                    dest_ws.cell(row=row, column=4).value = ""
                
                phone_no = ""
                if "PAYMENT" not in action_status:
                    raw_value = dest_ws.cell(row=row, column=12).value
                    if raw_value:
                        phone_no = str(raw_value).strip().split('.')[0]
                
                if "PTP NEW" in action_status:
                    phone_value = source_phone1 if source_phone1 else source_phone2
                    remark_text = f"1_{phone_value} - PTP NEW"
                elif "PTP FF" in action_status:
                    phone_value = source_phone1 if source_phone1 else source_phone2
                    remark_text = f"{phone_value} - FPTP"
                elif "PAYMENT" in action_status:
                    remark_text = "CURED - CONFIRM VIA SELECTIVE LIST"
                else:
                    remark_text = ""
                
                dest_ws.cell(row=row, column=7).value = remark_text
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=9).value = ""
                else:
                    dest_ws.cell(row=row, column=9).value = source_amount
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=10).value = source_amount
                else:
                    dest_ws.cell(row=row, column=10).value = ""
                
                dest_ws.cell(row=row, column=11).value = source_collector
                
                if "PAYMENT" in action_status:
                    dest_ws.cell(row=row, column=12).value = ""
                else:
                    if source_phone1 and source_phone1 != "":
                        dest_ws.cell(row=row, column=12).value = source_phone1
                    else:
                        dest_ws.cell(row=row, column=12).value = source_phone2
                
                if "PAYMENT" in action_status and source_date:
                    if isinstance(source_date, datetime):
                        formatted_paid_date = source_date.strftime("%m/%d/%Y")
                    elif isinstance(source_date, date):
                        formatted_paid_date = source_date.strftime("%m/%d/%Y")
                    else:
                        try:
                            date_obj = datetime.strptime(str(source_date), "%Y-%m-%d %H:%M:%S")
                            formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                        except:
                            try:
                                date_obj = datetime.strptime(str(source_date), "%Y-%m-%d")
                                formatted_paid_date = date_obj.strftime("%m/%d/%Y")
                            except:
                                formatted_paid_date = ""
                    dest_ws.cell(row=row, column=14).value = formatted_paid_date
                else:
                    dest_ws.cell(row=row, column=14).value = ""
            
            for row in range(2, final_row_count + 1):
                action_status = dest_ws.cell(row=row, column=2).value
                phone_no = dest_ws.cell(row=row, column=12).value
                
                if "PTP NEW" in action_status and phone_no:
                    dest_ws.cell(row=row, column=7).value = f"1_{phone_no} - PTP NEW"
                elif "PTP FF" in action_status and phone_no:
                    dest_ws.cell(row=row, column=7).value = f"{phone_no} - FPTP"
            
            for column in dest_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                dest_ws.column_dimensions[column_letter].width = adjusted_width

            for row_idx in range(2, dest_ws.max_row + 1):  
                for col_idx in [3, 4, 14]: 
                    cell = dest_ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell_value_str = str(cell.value)
                        cell.value = cell_value_str
                        cell.number_format = '@'

            dest_wb.save(remarks_path)
            
            others_wb = openpyxl.Workbook()
            others_ws = others_wb.active
            
            others_ws.cell(row=1, column=1).value = ws.cell(row=1, column=1).value 
            others_ws.cell(row=1, column=2).value = "REMARK BY" 
            
            for row in range(2, last_row + 1):
                others_ws.cell(row=row, column=1).value = ws.cell(row=row, column=1).value

                reference_value = ws.cell(row=row, column=1).value 
                
                for cured_row in range(2, ws.max_row + 1):  
                    if ws.cell(row=cured_row, column=1).value == reference_value: 
                        others_ws.cell(row=row, column=2).value = ws.cell(row=cured_row, column=2).value 
                        break

            others_wb.save(others_path)
            
            payments_wb = openpyxl.Workbook()
            payments_ws = payments_wb.active
            payments_ws.cell(row=1, column=1).value = "LAN"
            payments_ws.cell(row=1, column=2).value = "ACCOUNT NUMBER"
            payments_ws.cell(row=1, column=3).value = "NAME"
            payments_ws.cell(row=1, column=4).value = "CARD NUMBER"
            payments_ws.cell(row=1, column=5).value = "PAYMENT AMOUNT"
            payments_ws.cell(row=1, column=6).value = "PAYMENT DATE"
            
            for row in range(2, last_row + 1):
                payments_ws.cell(row=row, column=1).value = ws.cell(row=row, column=17).value if ws.cell(row=row, column=17).value else ""
                payments_ws.cell(row=row, column=3).value = ws.cell(row=row, column=18).value if ws.cell(row=row, column=18).value else ""
                payments_ws.cell(row=row, column=5).value = ws.cell(row=row, column=4).value if ws.cell(row=row, column=4).value else ""
                date_value = ws.cell(row=row, column=3).value
                if date_value:
                    if isinstance(date_value, datetime):
                        formatted_date = date_value.strftime("%m/%d/%Y")
                    else:
                        formatted_date = str(date_value)
                    payments_ws.cell(row=row, column=6).value = formatted_date
            
            for row in range(2, last_row + 1):
                payments_ws.cell(row=row, column=6).number_format = "@"
            
            for column in payments_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                payments_ws.column_dimensions[column_letter].width = adjusted_width
            
            payments_wb.save(payments_path)
            
            remarks_df = pd.read_excel(remarks_path)
            others_df = pd.read_excel(others_path)
            payments_df = pd.read_excel(payments_path)
            
            with open(remarks_path, 'rb') as f:
                remarks_binary = f.read()
            with open(others_path, 'rb') as f:
                others_binary = f.read()
            with open(payments_path, 'rb') as f:
                payments_binary = f.read()
                
            os.unlink(temp_input_path)
            
            return {
                'remarks_df': remarks_df, 
                'others_df': others_df, 
                'payments_df': payments_df,
                'remarks_binary': remarks_binary,
                'others_binary': others_binary,
                'payments_binary': payments_binary,
                'remarks_filename': remarks_filename,
                'others_filename': others_filename,
                'payments_filename': payments_filename
            }
        finally:
            if os.path.exists(temp_input_path):
                os.unlink(temp_input_path)
                
class ROBBikeProcessor(BaseProcessor):
    def process_daily_remark(self, file_content, sheet_name=None, preview_only=False,
                    remove_duplicates=False, remove_blanks=False, trim_spaces=False, report_date=None):
        try:
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            required_columns = ['Time', 'Status', 'Account No.', 'Debtor', 'DPD', 'Remark', 'Remark By', 'PTP Amount', 'Balance', 'Claim Paid Amount']
            
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                st.error("Required columns not found in the uploaded file.")
                return None, None, None
            else: 
                if 'Time' in df.columns:
                    if pd.api.types.is_object_dtype(df['Time']):
                        try:
                            df['Time'] = pd.to_datetime(df['Time'], format='%I:%M:%S %p')
                        except ValueError:
                            pass
                    df = df.sort_values(by='Time', ascending=False)
                
                if 'Status' in df.columns:
                    df['Status'] = df['Status'].fillna('')
                
                    dnc_mask = df['Status'].str.contains('DNC', case=False)
                    blank_mask = df['Status'].str.strip() == ''
                
                    removed_dnc_count = dnc_mask.sum()
                    removed_blank_count = blank_mask.sum()
                
                    df = df[~(dnc_mask | blank_mask)]
                    
                    disposition = supabase.table('rob_bike_disposition').select("disposition").execute()
                
                    if disposition.data is None:
                        valid_dispo = []
                    else:
                        valid_dispo = [record['disposition'] for record in disposition.data]
                
                    not_in_valid_dispo = ~df['Status'].isin(valid_dispo)
                    removed_invalid_dispo_count = not_in_valid_dispo.sum()
                    df = df[~not_in_valid_dispo]
                    
                if 'Account No.' in df.columns and 'Status' in df.columns:
                    initial_duplicates = df.duplicated(subset=['Account No.', 'Status']).sum()
                    df['COMBINED_KEY'] = df['Account No.'].astype(str) + '_' + df['Status'].astype(str)
                    #remaining_duplicates = df.duplicated(subset=['COMBINED_KEY']).sum()
                    df = df.drop_duplicates(subset=['COMBINED_KEY'])
                    df = df.drop(columns=['COMBINED_KEY'])
                
                if 'Remark' in df.columns:
                    system_auto_update_remarks = df['Remark'].str.contains('System Auto Update Remarks For PD', case=False, na=False)
                    system_auto_update_remarks_count = system_auto_update_remarks.sum()
                    df = df[~system_auto_update_remarks]
                
                if 'Remark By' in df.columns:
                    jerivera_remarks = df['Remark By'].str.contains('JERIVERA', case=False, na=False)
                    system_remarks_count = jerivera_remarks.sum()
                    df = df[~jerivera_remarks]
                    system_remarks = df['Remark By'].str.contains('SYSTEM', case=False, na=False)
                    system_remarks_count = system_remarks.sum()
                    df = df[~system_remarks]
                    
                if 'PTP Amount' in df.columns and 'Balance' in df.columns and 'Claim Paid Amount' in df.columns:
                    df['PTP Amount'] = pd.to_numeric(df['PTP Amount'].replace({',': ''}, regex=True), errors='coerce')
                    df['Balance'] = pd.to_numeric(df['Balance'].replace({',': ''}, regex=True), errors='coerce')
                    df['Claim Paid Amount'] = pd.to_numeric(df['Claim Paid Amount'].replace({',': ''}, regex=True), errors='coerce')
                
                if 'PTP Amount' in df.columns and 'Status' in df.columns:
                    voluntary_surrender_rows = df[df['Status'] == 'PTP - VOLUNTARY SURRENDER']

                    invalid_amount_rows = voluntary_surrender_rows[
                        (voluntary_surrender_rows['PTP Amount'].isna()) |
                        (voluntary_surrender_rows['PTP Amount'] == 0)
                    ]

                    if not invalid_amount_rows.empty:
                        st.warning(f"Found {len(invalid_amount_rows)} row(s) with 'PTP - VOLUNTARY SURRENDER' but 0 or missing 'PTP Amount'.")
                        st.dataframe(invalid_amount_rows, use_container_width=True)
                        
                st.write(f"Removed: {removed_dnc_count} DNC, {removed_blank_count} blank status, {removed_invalid_dispo_count} invalid disposition, {system_auto_update_remarks_count} system auto update remarks, {system_remarks_count} system remarks, {initial_duplicates} duplicates.")

                if preview_only:
                    return df, None, None
                
                output_template = "DAILY MONITORING PTP, DEPO & REPO REPORT TEMPLATE.xlsx"
                sheet1 = "MONITORING"
                sheet2 = "PTP"
                sheet3 = "REPO"
                sheet4 = "DEPO"
                sheet5 = "EOD"
                
                monitoring_columns = ['Account Name', 'Account Number', 'Principal', 'EndoDate', 'Stores', 
                                    'Cluster', 'DaysPastDue', 'Field Status', 'Field Substatus', 
                                    'Status', 'subStatus', 'Notes', 'BarcodeDate', 'PTP Amount', 'PTP Date']
                monitoring_df = pd.DataFrame(columns=monitoring_columns)
                
                ptp_columns = ['Account Name', 'AccountNumber', 'Status', 'subStatus', 'Amount', 
                            'StartDate', 'Notes', 'ResultDate', 'EndoDate']
                ptp_df = pd.DataFrame(columns=ptp_columns)
                
                eod_df = pd.DataFrame()
                
                if 'Debtor' in df.columns:
                    monitoring_df['Account Name'] = df['Debtor'].str.upper()
                
                if 'Account No.' in df.columns:
                    monitoring_df['Account Number'] = df['Account No.']
                
                if 'Balance' in df.columns:
                    monitoring_df['Principal'] = df['Balance']
                
                if 'DPD' in df.columns:
                    monitoring_df['DaysPastDue'] = df['DPD']
                
                if 'Status' in df.columns:
                    status_parts = df['Status'].str.split('-', n=1)
                    monitoring_df['Status'] = status_parts.str[0].str.strip()
                    monitoring_df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
                
                if 'Remark' in df.columns:
                    monitoring_df['Notes'] = df['Remark']
                
                if 'Date' in df.columns:
                    monitoring_df['BarcodeDate'] = pd.to_datetime(df['Date']).dt.strftime('%m/%d/%Y')
                
                if 'PTP Amount' in df.columns and 'Claim Paid Amount' in df.columns:
                    ptp_amount = df['PTP Amount']
                    ptp_date = pd.to_datetime(df['PTP Date'], errors='coerce')
                    claim_paid_amount = df['Claim Paid Amount']
                    claim_paid_date = pd.to_datetime(df['Claim Paid Date'], errors='coerce')
                    
                    monitoring_df['PTP Amount'] = np.where(
                        ptp_amount.notna() & (ptp_amount != 0),
                        ptp_amount,
                        np.where(
                            claim_paid_amount.notna() & (claim_paid_amount != 0),
                            claim_paid_amount,
                            ''
                        )
                    )
                    
                    monitoring_df['PTP Date'] = np.where(
                        ptp_date.notna(),
                        ptp_date.dt.strftime('%m/%d/%Y'),
                        np.where(
                            claim_paid_date.notna(),
                            claim_paid_date.dt.strftime('%m/%d/%Y'),
                            ''
                        )
                    )
                                    
                if 'Account No.' in df.columns:
                    account_numbers = [str(int(acc)) for acc in df['Account No.'].dropna().unique().tolist()]
                    dataset_response = supabase.table('rob_bike_dataset').select('*').in_('account_number', account_numbers).execute()
                    
                    if hasattr(dataset_response, 'data') and dataset_response.data:
                        dataset_df = pd.DataFrame(dataset_response.data)
                        monitoring_df['Account Number'] = monitoring_df['Account Number'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
                        
                        account_data_map = {}
                        chcode_list = []
                        
                        for _, row in dataset_df.iterrows():
                            account_no = str(row['account_number']).strip()
                            chcode = row.get('chcode', '')
                            
                            if chcode:
                                chcode_list.append(chcode)
                                
                            account_data_map[account_no] = {
                                'ChCode': chcode,
                                'AccountNumber': "00" + account_no,
                                'EndoDate': row.get('endo_date', ''),
                                'Stores': row.get('stores', ''),
                                'Cluster': row.get('cluster', '')
                            }
                        
                        if chcode_list:
                            try:
                                field_results_response = supabase.table('rob_bike_field_result').select('*').in_('chcode', chcode_list).execute()
                                
                                if hasattr(field_results_response, 'data') and field_results_response.data:
                                    field_results_df = pd.DataFrame(field_results_response.data)
                                    
                                    if 'inserted_date' in field_results_df.columns:
                                        field_results_df['inserted_date'] = pd.to_datetime(field_results_df['inserted_date'])
                                    
                                    latest_status_map = {}
                                    
                                    if 'inserted_date' in field_results_df.columns:
                                        for chcode, group in field_results_df.groupby('chcode'):
                                            latest_row = group.sort_values('inserted_date', ascending=False).iloc[0]
                                            
                                            status = latest_row.get('status', '')
                                            substatus = latest_row.get('substatus', '')
                                            
                                            if status in ('0', '') or substatus in ('0', ''):
                                                status, substatus = '', ''
                                            
                                            latest_status_map[chcode] = {
                                                'Field_Status': status if status not in ('0', '') else '',
                                                'Field_Substatus': substatus if substatus not in ('0', '') else '',
                                            }
                                        
                                        for account_no, data in account_data_map.items():
                                            chcode = data['ChCode']
                                            if chcode in latest_status_map:
                                                account_data_map[account_no].update({
                                                    'Field_Status': latest_status_map[chcode]['Field_Status'],
                                                    'Field_Substatus': latest_status_map[chcode]['Field_Substatus'],
                                                })
                                            else:
                                                account_data_map[account_no].update({
                                                    'Field_Status': '',
                                                    'Field_Substatus': '',
                                                })
                                            
                            except Exception as e:
                                st.error(f"Error fetching field results: {str(e)}")
                        
                        monitoring_df['EndoDate'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('EndoDate', ''))
                        monitoring_df['EndoDate'] = pd.to_datetime(monitoring_df['EndoDate']).dt.strftime('%m/%d/%Y')
                        
                        monitoring_df['Stores'] = monitoring_df['Account Number'].map(
                            lambda acc_no: '' if account_data_map.get(acc_no, {}).get('Stores') in ['0', 0] 
                            else account_data_map.get(acc_no, {}).get('Stores', '')
                        )
                        
                        monitoring_df['Cluster'] = monitoring_df['Account Number'].map(
                            lambda acc_no: '' if account_data_map.get(acc_no, {}).get('Cluster') in ['0', 0] 
                            else account_data_map.get(acc_no, {}).get('Cluster', '')
                        )
                        
                        monitoring_df['Field Status'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('Field_Status', ''))
                        
                        monitoring_df['Field Substatus'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('Field_Substatus', ''))
                        
                        monitoring_df['Account Number'] = monitoring_df['Account Number'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('AccountNumber', ''))
                        
                ptp_data = df[df['Status'].str.contains('PTP', case=False, na=False)].copy() if 'Status' in df.columns else pd.DataFrame()
                
                if not ptp_data.empty:
                    if 'Debtor' in ptp_data.columns:
                        ptp_df['Account Name'] = ptp_data['Debtor'].str.upper()
                    
                    if 'Account No.' in ptp_data.columns:
                        ptp_df['AccountNumber'] = ptp_data['Account No.']
                    
                    if 'Status' in ptp_data.columns:
                        status_parts = ptp_data['Status'].str.split('-', n=1)
                        ptp_df['Status'] = status_parts.str[0].str.strip()
                        ptp_df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
                    
                    if 'PTP Amount' in ptp_data.columns:
                        ptp_df['Amount'] = ptp_data['PTP Amount']
                    
                    if 'PTP Date' in ptp_data.columns:
                        ptp_df['StartDate'] = pd.to_datetime(ptp_data['PTP Date']).dt.strftime('%Y-%m-%d')
                    
                    if 'Remark' in ptp_data.columns:
                        ptp_df['Notes'] = ptp_data['Remark']
                    
                    if 'Time' in ptp_data.columns:
                        time_only = pd.to_datetime(ptp_data['Time'], errors='coerce').dt.time

                        result_datetime = [
                            datetime.combine(report_date, t) if pd.notnull(t) else None for t in time_only
                        ]

                        ptp_df['ResultDate'] = [
                            dt.strftime('%m/%d/%Y %I:%M:%S %p').replace(' 0', ' ') if dt else '' for dt in result_datetime
                        ]
                        
                    if 'Account No.' in ptp_data.columns and 'account_data_map' in locals():
                        ptp_df['AccountNumber'] = ptp_df['AccountNumber'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else '')
                        ptp_df['EndoDate'] = ptp_df['AccountNumber'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('EndoDate', ''))
                        ptp_df['EndoDate'] = pd.to_datetime(ptp_df['EndoDate']).dt.strftime('%m/%d/%Y')
                
                    if 'Account No.' in df.columns:
                        ptp_df['AccountNumber'] = ptp_df['AccountNumber'].map(
                            lambda acc_no: account_data_map.get(acc_no, {}).get('AccountNumber', ''))
            
                payment_statuses = [
                    "PAYMENT", "PAYMENT VIA CALL", "PAYMENT VIA SMS", "PAYMENT VIA EMAIL",
                    "PAYMENT VIA FIELD VISIT", "PAYMENT VIA CARAVAN", "PAYMENT VIA SOCMED"
                ]
                ptp_statuses = [
                    "PTP", "PTP VIA CALL", "PTP VIA SMS", "PTP VIA EMAIL", "PTP VIA FIELD VISIT",
                    "PTP VIA CARAVAN", "PTP VIA SOCMED"
                ]
                if 'Status' in df.columns:
                    status_parts = df['Status'].str.split('-', n=1)
                    df['Status'] = status_parts.str[0].str.strip()
                    
                    df['subStatus'] = status_parts.str[1].str.strip().where(status_parts.str.len() > 1, "")
                    
                df['Status'] = df['Status'].astype(str)
                df['subStatus'] = df['subStatus'].astype(str)
                
                total_principal = df['Balance'].sum()
                total_accounts = df['Balance'].count()
                
                filtered_vs = df[
                    (df['Status'].isin(payment_statuses)) &
                    (df['subStatus'].str.upper() == "VOLUNTARY SURRENDER")
                ]   
                repo_amount = filtered_vs['Balance'].sum()
                repo_count = filtered_vs['Balance'].count()

                filtered_payment = df[
                    (df['Status'].isin(ptp_statuses)) &
                    (~df['subStatus'].str.contains("Follow up", case=False, na=False))
                ]
                ptp_amount = filtered_payment['Balance'].sum()
                
                filtered_ptp = df[
                    (df['Status'].str.contains("PTP", case=False, na=False)) &
                    (~df['subStatus'].str.contains("Follow up", case=False, na=False))
                ]
                ptp_count = filtered_ptp.shape[0]
                
                eod_data = {
                    'Key': ['C2', 'D2', 'C5', 'D5', 'C9', 'D9'],
                    'Value': [total_principal, total_accounts, repo_amount, repo_count, ptp_amount, ptp_count]
                }
                eod_df = pd.DataFrame(eod_data)

                priority_substatus = [
                    ("FULLY PAID", "PAY OFF"),
                    ("PARTIAL", "STILL PD BUT WITH ARRANGEMENT"),
                    ("FULL UPDATE", "CURRENT")
                ]

                bottom_rows = []
                row_index = 12

                for substatus_value, label in priority_substatus:
                    temp_df = df[
                        (df['Status'].isin(payment_statuses)) &
                        (df['subStatus'].str.upper().str.contains(substatus_value.upper()))
                    ]
                    
                    for _, row in temp_df.iterrows():
                        bottom_rows.append({
                            'Key': f'C{row_index}',
                            'Value': row['Balance']
                        })
                        
                        ptp_value = row['Claim Paid Amount']
                        if ptp_value == 0 or ptp_value == '':
                            ptp_value = row['PTP Amount']
                            
                        bottom_rows.append({
                            'Key': f'D{row_index}',
                            'Value': ptp_value
                        })
                        
                        bottom_rows.append({
                            'Key': f'E{row_index}',
                            'Value': label
                        })
                        
                        row_index += 1

                min_rows = 2
                end_row = max(row_index, 12 + min_rows)
                for blank_row in range(row_index, end_row):
                    bottom_rows.append({'Key': f'C{blank_row}', 'Value': ''})
                    bottom_rows.append({'Key': f'E{blank_row}', 'Value': ''})
                    
                eod_df = pd.concat([eod_df, pd.DataFrame(bottom_rows)], ignore_index=True)
                
                template_path = os.path.join(os.path.dirname(__file__), "templates", "rob_bike", output_template)
                
                output_buffer = io.BytesIO()
                
                if os.path.exists(template_path):
                    try:
                        with open(template_path, 'rb') as template_file:
                            template_copy = io.BytesIO(template_file.read())
                            
                        try:
                            template_wb = load_workbook(template_copy)
                            
                            def append_df_to_sheet(sheet_name, df):
                                if sheet_name in template_wb.sheetnames:
                                    sheet = template_wb[sheet_name]
                                    start_row = sheet.max_row + 1
                                    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row):
                                        for c_idx, value in enumerate(row, 1):
                                            sheet.cell(row=r_idx, column=c_idx).value = value
                            
                            append_df_to_sheet(sheet1, monitoring_df)
                            append_df_to_sheet(sheet2, ptp_df)
                            
                            def format_sheet(sheet_name, df=None):
                                sheet = template_wb[sheet_name]
                                
                                thin_border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'),
                                )

                                if df is not None:
                                    for col_idx, col in enumerate(df.columns, 1):
                                        max_length = max(
                                            df[col].astype(str).map(len).max(),
                                            len(str(col))
                                        )
                                        adjusted_width = max_length + 2
                                        sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                                    start_row = sheet.max_row - len(df) + 1
                                    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=1, max_col=len(df.columns)):
                                        for cell in row:
                                            cell.border = thin_border
                                            
                            if sheet5 in template_wb.sheetnames:
                                eod_sheet = template_wb[sheet5]
                                for _, row in eod_df.iterrows():
                                    cell_key = row['Key']
                                    value = row['Value']
                                    column_letter = cell_key[0]
                                    row_number = int(cell_key[1:])
                                    column_index = column_index_from_string(column_letter)
                                    eod_sheet.cell(row=row_number, column=column_index).value = value
                            
                            format_sheet(sheet1, monitoring_df)
                            format_sheet(sheet2, ptp_df)
                            format_sheet(sheet5, None)
                            
                            template_wb.save(output_buffer)
                            
                        except Exception as e:
                            st.error(f"Error processing template: {str(e)}")
                            
                    except Exception as e:
                        st.error(f"Error reading template file: {str(e)}")
                    
                else:
                    st.write("Template does not exist")
                    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                        monitoring_df.to_excel(writer, sheet_name=sheet1, index=False)
                        ptp_df.to_excel(writer, sheet_name=sheet2, index=False)
                        eod_df.to_excel(writer, sheet_name=sheet5, index=False)

                        workbook = writer.book
                        
                        workbook.create_sheet(title=sheet3)
                        workbook.create_sheet(title=sheet4)
                        
                        def format_sheet(sheet_name, df=None):
                            sheet = writer.sheets.get(sheet_name) or workbook[sheet_name]
                            
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'),
                            )
                            
                            if df is not None: 
                                for col_idx, col in enumerate(df.columns, 1):
                                    max_length = max(
                                        df[col].astype(str).map(len).max(),
                                        len(str(col))
                                    )
                                    adjusted_width = max_length + 2
                                    sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                                
                                for row in sheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(df.columns)):
                                    for cell in row:
                                        cell.border = thin_border
                                        
                        eod_sheet = workbook[sheet5]
                        for _, row in eod_df.iterrows():
                            cell_key = row['Key']
                            value = row['Value']
                            column_letter = cell_key[0]
                            row_number = int(cell_key[1:])
                            column_index = column_index_from_string(column_letter)
                            eod_sheet.cell(row=row_number, column=column_index).value = value
                            
                        format_sheet(sheet1, monitoring_df)
                        format_sheet(sheet2, ptp_df)
                        format_sheet(sheet3)
                        format_sheet(sheet4)
                        format_sheet(sheet5, None)
                        
                output_buffer.seek(0)
                
                if not report_date:
                    report_date = datetime.now()

                date_str = report_date.strftime("%d%b%Y").upper()
                
                output_filename = f"DAILY MONITORING PTP, DEPO & REPO REPORT as of {date_str}.xlsx"
                
                return monitoring_df, output_buffer.getvalue(), output_filename
        
        except Exception as e:
            st.error(f"Error processing daily remark: {str(e)}")
            return None, None, None

    def process_new_endorsement(self, file_content, sheet_name=None, preview_only=False,
                            remove_duplicates=False, remove_blanks=False, trim_spaces=False):
        try:
            if isinstance(file_content, bytes):
                file_content = io.BytesIO(file_content)
            
            xls = pd.ExcelFile(file_content)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)
            
            if 'Endorsement Date' in df.columns:
                df = df.drop(columns='Endorsement Date')
            
            if 'Account Number' in df.columns:
                account_numbers_list = [str(acc) for acc in df['Account Number'].dropna().unique().tolist()]
                
                batch_size = 100 
                existing_accounts = []
                
                for i in range(0, len(account_numbers_list), batch_size):
                    batch = account_numbers_list[i:i + batch_size]
                    response = supabase.table('rob_bike_dataset').select('account_number').in_('account_number', batch).execute()
                    
                    if hasattr(response, 'data') and response.data:
                        existing_accounts.extend([str(item['account_number']) for item in response.data])
                
                initial_rows = len(df)
                df = df[~df['Account Number'].astype(str).isin(existing_accounts)]
                removed_rows = initial_rows - len(df)
                
                if removed_rows > 0:
                    st.write(f"Removed {removed_rows} rows with existing account numbers")
                
                if df.empty:
                    st.warning("No new account numbers found (all account numbers exists)")
                    return None, None, None
            
            current_date = datetime.now().strftime('%Y/%m/%d')
            df.insert(0, 'ENDO DATE', current_date)
            
            if 'Endrosement OB' in df.columns:
                df['Endrosement OB'] = pd.to_numeric(df['Endrosement OB'], errors='coerce')
                zero_ob_rows = df[df['Endrosement OB'] == 0]
                if not zero_ob_rows.empty:
                    st.warning(f"Found {len(zero_ob_rows)} rows with 0 in Endorsement OB")
            
            if preview_only:
                return df, None, None
            
            result_df = df
            output_filename = f"rob_bike-new-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
            output_path = os.path.join(os.getcwd(), output_filename)  
            
            # numeric_cols = ['Endrosement OB']  
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                result_df.to_excel(writer, index=False, sheet_name='Sheet1')
                
                worksheet = writer.sheets['Sheet1']
                final_columns = result_df.columns
                
                for i, col in enumerate(final_columns):
                    col_letter = chr(65 + i)
                    
                    # if col in numeric_cols:
                    #     for row in range(2, len(result_df) + 2):
                    #         cell = worksheet[f"{col_letter}{row}"]
                    #         cell.number_format = '0.00'
                    
                    if col == 'ENDO DATE':
                        for row in range(2, len(result_df) + 2):
                            cell = worksheet[f"{col_letter}{row}"]
                            value = cell.value
                            if value:
                                try:
                                    cell.value = pd.to_datetime(value).strftime("%m/%d/%Y")
                                    cell.number_format = '@'
                                except:
                                    pass
            
            with open(output_path, 'rb') as f:
                output_binary = f.read()
            
            return result_df, output_binary, output_filename
        
        except Exception as e:
            st.error(f"Error processing new endorsement: {str(e)}")
            return None, None, None

class BDOAutoProcessor(BaseProcessor):
    def process_agency_daily_report(self, file_content, sheet_name=None, preview_only=False,
        remove_duplicates=False, remove_blanks=False, trim_spaces=False, report_date=None,
        kept_count_b5=None, kept_bal_b5=None, alloc_bal_b5=None,
        kept_count_b6=None, kept_bal_b6=None, alloc_bal_b6=None):

        try:
            DIR = os.getcwd()
            
            TEMPLATE_DIR = os.path.join(DIR, "templates", "bdo_auto")
            daily_report_template = os.path.join(TEMPLATE_DIR, "AGENCY DAILY REPORT TEMPLATE.xlsx")
            daily_productivity_template = os.path.join(TEMPLATE_DIR, "DAILY PRODUCTIVITY TEMPLATE.xlsx")
            vs_report_template = os.path.join(TEMPLATE_DIR, "SPMADRID VS REPORT TEMPLATE.xlsx")
            
            if not os.path.exists(daily_report_template):
                st.error(f"Template file not found: {daily_report_template}")
                return None, None, None
                
            if not os.path.exists(daily_productivity_template):
                st.error(f"Template file not found: {daily_productivity_template}")
                return None, None, None
                
            try:
                with open(daily_report_template, 'rb') as template_file:
                    template_copy = io.BytesIO(template_file.read())
                try:
                    test_wb = load_workbook(template_copy)
                    test_wb.close()
                except zipfile.BadZipFile:
                    st.error(f"Template file is not a valid Excel file: {daily_report_template}")
                    return None, None, None
            except Exception as e:
                st.error(f"Error opening daily report template file: {str(e)}")
                return None, None, None
                
            try:
                with open(daily_productivity_template, 'rb') as template_file:
                    template_copy = io.BytesIO(template_file.read())
                try:
                    test_wb = load_workbook(template_copy)
                    test_wb.close()
                except zipfile.BadZipFile:
                    st.error(f"Template file is not a valid Excel file: {daily_productivity_template}")
                    return None, None, None
            except Exception as e:
                st.error(f"Error opening daily productivity template file: {str(e)}")
                return None, None, None
            
            BASE_DIR = os.path.join(DIR, "database", "bdo_auto")
            
            bucket_paths = {
                "Bucket 1": os.path.join(BASE_DIR, "BUCKET1_AGENT.xlsx"),
                "Bucket 2": os.path.join(BASE_DIR, "BUCKET2_AGENT.xlsx"),
                "Bucket 5&6": os.path.join(BASE_DIR, "BUCKET5&6_AGENT.xlsx")
            }
        
            bank_status_path = os.path.join(BASE_DIR, "BANK_STATUS.xlsx")
            rfd_list = os.path.join(BASE_DIR, "RFD_LISTS.xlsx")
            
            expected_columns = [
                "Date", "Debtor", "Account No.", "Card No.", "Remark", "Remark By",
                "PTP Amount", "PTP Date", "Claim Paid Amount", "Claim Paid Date", 
                "Balance", "Status"
            ]
            
            bank_status_lookup = {}
            if os.path.exists(bank_status_path):
                df_bank_status = pd.read_excel(bank_status_path)
                if "CMS STATUS" not in df_bank_status.columns or "BANK STATUS" not in df_bank_status.columns:
                    st.error("Missing 'CMS STATUS' or 'BANK STATUS' column in BANK_STATUS.xlsx.")
                    return None, None, None
                bank_status_lookup = dict(zip(df_bank_status["CMS STATUS"].astype(str).str.strip(), 
                                            df_bank_status["BANK STATUS"].astype(str).str.strip()))
            else:
                st.error(f"Missing file: {bank_status_path}")
                return None, None, None
                
            rfd_valid_codes = set()
            if os.path.exists(rfd_list):
                df_rfd_list = pd.read_excel(rfd_list)
                if "RFD CODE" not in df_rfd_list.columns:
                    st.error("Missing 'RFD CODE' column in RFD_LISTS.xlsx.")
                    return None, None, None
                rfd_valid_codes = set(df_rfd_list["RFD CODE"].astype(str).str.upper())
            else:
                st.error(f"Missing file: {rfd_list}")
                return None, None, None
                
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            df_main = pd.read_excel(xls, sheet_name=sheet_name, dtype={"Account No.": str})
            
            df_main = self.clean_data(df_main, remove_duplicates, remove_blanks, trim_spaces)
            
            missing_columns = [col for col in expected_columns if col not in df_main.columns]
            if missing_columns:
                st.error("Required columns not found in the uploaded file.")
                return None, None, None
                
            df_main["Remark By"] = df_main["Remark By"].astype(str).str.strip()
            
            df_main = df_main[~df_main["Remark"].isin([
                "Updates when case reassign to another collector", 
                "System Auto Update Remarks For PD"
            ])]
            
            df_main = df_main[~df_main["Card No."].isin([f"ch{i}" for i in range(1, 20)])]
            
            bucket_dfs = {}
            for bucket_name, bucket_path in bucket_paths.items():
                if os.path.exists(bucket_path):
                    df_bucket = pd.read_excel(bucket_path)
                    if "VOLARE USER" not in df_bucket.columns or "FULL NAME" not in df_bucket.columns:
                        st.warning(f"{bucket_name} missing required columns. Skipping.")
                        continue
                        
                    df_bucket["VOLARE USER"] = df_bucket["VOLARE USER"].astype(str).str.strip()
                    df_bucket["FULL NAME"] = df_bucket["FULL NAME"].astype(str).str.strip()
                    lookup_dict = dict(zip(df_bucket["VOLARE USER"], df_bucket["FULL NAME"]))
                    
                    matched_df = df_main[df_main["Remark By"].isin(df_bucket["VOLARE USER"])].copy()
                    matched_df["HANDLING OFFICER2"] = matched_df["Remark By"].map(lookup_dict)
                    
                    if bucket_name == "Bucket 1":
                        matched_df = matched_df[
                            (matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]) &
                            matched_df["Card No."].astype(str).str.startswith("01")) |
                            (~matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]))
                        ]
                    elif bucket_name == "Bucket 2":
                        matched_df = matched_df[
                            (matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]) &
                            matched_df["Card No."].astype(str).str.startswith("02")) |
                            (~matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]))
                        ]
                    elif bucket_name == "Bucket 5&6":
                        matched_df = matched_df[
                            (matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]) &
                            matched_df["Card No."].astype(str).str.startswith(("05", "06"))) |
                            (~matched_df["Remark By"].isin(["SYSTEM", "LCMANZANO", "ACALVAREZ", "DSDEGUZMAN", "SRELIOT", "TANAZAIRE", "SPMADRID"]))
                        ]
                    
                    for col in ["PTP Date", "Claim Paid Date", "Date"]:
                        matched_df[col] = pd.to_datetime(matched_df[col], errors='coerce')
                    
                    matched_df["BANK STATUS"] = matched_df["Status"].astype(str).str.strip().map(bank_status_lookup)
                    
                    if not matched_df.empty:
                        bucket_dfs[bucket_name] = matched_df
                else:
                    st.error(f"Missing file: {bucket_path}")
            
            def extract_and_validate_rfd(remark):
                remark = str(remark).strip().rstrip("\\")
                rfd_match = re.search(r"RFD:\s*(\S+)$", remark)
                if rfd_match:
                    rfd = rfd_match.group(1).upper()
                else:
                    last_word = re.findall(r"\\\s*(\S+)", remark)
                    if last_word:
                        rfd = last_word[-1].upper()
                    else:
                        last_word = remark.split()[-1] if remark else np.nan
                        rfd = last_word.upper() if last_word else np.nan
                return rfd if rfd in rfd_valid_codes else np.nan
            
            def autofit_worksheet_columns(ws):
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2
                    ws.column_dimensions[col_letter].width = adjusted_width
            
            def get_merged_cell_top_left(ws, cell_ref):
                """Find the top-left cell of a merged range containing the given cell_ref."""
                for merged_range in ws.merged_cells.ranges:
                    if cell_ref in merged_range:
                        return merged_range.min_row, merged_range.min_col
                return None, None  
            
            processed_dfs = {}
            for bucket_name, bucket_df in bucket_dfs.items():
                filtered_df = pd.DataFrame({
                    "Card Number": bucket_df["Card No."],
                    "PN": bucket_df["Account No."].astype(str).str.replace(r'\.0$', '', regex=True),
                    "NAME": bucket_df["Debtor"],
                    "BALANCE": bucket_df["Balance"].replace({',': ''}, regex=True).astype(float),
                    "HANDLING OFFICER2": bucket_df["HANDLING OFFICER2"].str.upper(),
                    "AGENCY3": "SP MADRID",
                    "STATUS4": bucket_df["BANK STATUS"],
                    "DATE OF CALL": bucket_df["Date"].dt.strftime("%m/%d/%Y"),
                    "PTP DATE": np.where(
                        bucket_df["PTP Date"].isna(),
                        np.where(bucket_df["Claim Paid Date"].isna(), np.nan, bucket_df["Claim Paid Date"].dt.strftime("%m/%d/%Y")),
                        bucket_df["PTP Date"].dt.strftime("%m/%d/%Y")
                    ),
                    "PTP AMOUNT": np.where(
                        bucket_df["PTP Amount"].isna() | (bucket_df["PTP Amount"] == 0),
                        np.where(bucket_df["Claim Paid Amount"].isna() | (bucket_df["Claim Paid Amount"] == 0), np.nan, bucket_df["Claim Paid Amount"]),
                        bucket_df["PTP Amount"]
                    ),
                    "RFD5": bucket_df["Remark"].apply(extract_and_validate_rfd)
                })
                
                filtered_df.reset_index(drop=True, inplace=True)
                for i in range(1, len(filtered_df)):
                    if filtered_df.loc[i, "HANDLING OFFICER2"] == "SYSTEM":
                        filtered_df.loc[i, "HANDLING OFFICER2"] = filtered_df.loc[i-1, "HANDLING OFFICER2"]
                
                filtered_df.loc[filtered_df["RFD5"].isna() & (filtered_df["STATUS4"] == "PTP"), "RFD5"] = "BUSY"
                filtered_df.loc[filtered_df["RFD5"].isna() & (filtered_df["STATUS4"] == "CALL NO PTP"), "RFD5"] = "NISV"
                filtered_df.loc[filtered_df["RFD5"].isna() & (filtered_df["STATUS4"] == "UNCON"), "RFD5"] = "NABZ"
                
                filtered_df = filtered_df[~(filtered_df["STATUS4"].isna() | (filtered_df["STATUS4"] == "EXCLUDE"))]
                
                filtered_df.loc[filtered_df["STATUS4"] != "PTP", "PTP DATE"] = np.nan
                filtered_df.loc[filtered_df["STATUS4"] != "PTP", "PTP AMOUNT"] = np.nan
                
                processed_dfs[bucket_name] = filtered_df
            
            if preview_only:
                preview_data = {}
                for bucket_name, filtered_df in processed_dfs.items():
                    preview_df = filtered_df.drop(columns=["Card Number"])
                    preview_data[bucket_name] = preview_df.head(10)
                return preview_data, len(df_main), None

            bucket_5_6_df = processed_dfs.get("Bucket 5&6", pd.DataFrame())
            
            if not bucket_5_6_df.empty:
                bucket5_df = bucket_5_6_df[bucket_5_6_df["Card Number"].astype(str).str.startswith("05")].copy()
                bucket6_df = bucket_5_6_df[bucket_5_6_df["Card Number"].astype(str).str.startswith("06")].copy()
                
                bucket5_df = bucket5_df.drop(columns=["Card Number"])
                bucket6_df = bucket6_df.drop(columns=["Card Number"])
                
                if not report_date:
                    day = datetime.now().day
                    month = datetime.now().strftime("%B")
                    current_date = f"{month} {day}".upper()
                else:
                    current_date = report_date
                current_date_formatted = datetime.now().strftime("%m/%d/%Y") if not report_date else datetime.strptime(report_date, "%B %d").strftime("%m/%d/%Y")

                if current_date.endswith(" 0"):
                    current_date = current_date[:-2] + current_date[-1:]
                
                output_files = {}
                productivity_files = {}
                b5_prod_df = None
                b6_prod_df = None
                
                template_wb = load_workbook(daily_report_template)
                
                if not bucket5_df.empty:
                    wb5 = load_workbook(daily_report_template)
                    ws5 = wb5.active
                    
                    headers = bucket5_df.columns.tolist()
                    for col_idx, header in enumerate(headers, 1):
                        ws5.cell(row=1, column=col_idx, value=header)
                    
                    for r_idx, row in enumerate(bucket5_df.values, 2):
                        for c_idx, value in enumerate(row, 1):
                            ws5.cell(row=r_idx, column=c_idx, value=value)
                    
                    autofit_worksheet_columns(ws5)
                    
                    output_b5 = io.BytesIO()
                    wb5.save(output_b5)
                    output_b5.seek(0)
                    b5_binary = output_b5
                    output_files["B5"] = b5_binary.getvalue()
                    
                    wb5_prod = load_workbook(daily_productivity_template)
                    ws5_prod = wb5_prod.active
                    
                    row, col = get_merged_cell_top_left(ws5_prod, 'C2')
                    if row and col:
                        ws5_prod.cell(row=row, column=col, value=current_date_formatted)
                    else:
                        ws5_prod['C2'] = current_date_formatted
                    
                    ptp_rows_b5 = bucket5_df[bucket5_df["STATUS4"] == "PTP"]
                    ptp_count_b5 = len(ptp_rows_b5)
                    ptp_balance_sum_b5 = ptp_rows_b5["BALANCE"].sum() if ptp_count_b5 > 0 else 0.0
                    
                    b5_prod_df = pd.DataFrame({
                        "Date": [current_date_formatted],
                        "PTP Count": [ptp_count_b5],
                        "Balance Sum": [ptp_balance_sum_b5],
                        "Kept Count": [kept_count_b5],
                        "Kept Balance": [kept_bal_b5],
                        "Allocation Balance": [alloc_bal_b5]
                    })
                    
                    ws5_prod['F8'] = ptp_count_b5
                    ws5_prod['G8'] = ptp_balance_sum_b5
                    ws5_prod['G8'].number_format = "0.00"
                    ws5_prod["K8"] = kept_count_b5
                    ws5_prod["K9"] = kept_count_b5
                    ws5_prod["L8"] = kept_bal_b5
                    ws5_prod["C13"] = alloc_bal_b5

                    autofit_worksheet_columns(ws5_prod)
                    
                    output_b5_prod = io.BytesIO()
                    wb5_prod.save(output_b5_prod)
                    output_b5_prod.seek(0)
                    productivity_files["B5"] = output_b5_prod.getvalue()
                    
                if not bucket6_df.empty:
                    wb6 = load_workbook(daily_report_template)
                    ws6 = wb6.active
                    
                    headers = bucket6_df.columns.tolist()
                    for col_idx, header in enumerate(headers, 1):
                        ws6.cell(row=1, column=col_idx, value=header)
                    
                    for r_idx, row in enumerate(bucket6_df.values, 2):
                        for c_idx, value in enumerate(row, 1):
                            ws6.cell(row=r_idx, column=c_idx, value=value)
                    
                    autofit_worksheet_columns(ws6)
                    
                    output_b6 = io.BytesIO()
                    wb6.save(output_b6)
                    output_b6.seek(0)
                    b6_binary = output_b6
                    output_files["B6"] = b6_binary.getvalue()
                    
                    wb6_prod = load_workbook(daily_productivity_template)
                    ws6_prod = wb6_prod.active
                    
                    row, col = get_merged_cell_top_left(ws6_prod, 'C2')
                    if row and col:
                        ws6_prod.cell(row=row, column=col, value=current_date_formatted)
                    else:
                        ws6_prod['C2'] = current_date_formatted
                    
                    ptp_rows_b6 = bucket6_df[bucket6_df["STATUS4"] == "PTP"]
                    ptp_count_b6 = len(ptp_rows_b6)
                    ptp_balance_sum_b6 = ptp_rows_b6["BALANCE"].sum() if ptp_count_b6 > 0 else 0.0
                    
                    b6_prod_df = pd.DataFrame({
                        "Date": [current_date_formatted],
                        "PTP Count": [ptp_count_b6],
                        "Balance Sum": [ptp_balance_sum_b6],
                        "Kept Count": [kept_count_b6],
                        "Kept Balance": [kept_bal_b6],
                        "Allocation Balance": [alloc_bal_b6]
                    })
                    
                    ws6_prod['F8'] = ptp_count_b6
                    ws6_prod['G8'] = ptp_balance_sum_b6
                    ws6_prod['G8'].number_format = "0.00"
                    ws6_prod["K8"] = kept_count_b6
                    ws6_prod["K9"] = kept_count_b6
                    ws6_prod["L8"] = kept_bal_b6
                    ws6_prod["C13"] = alloc_bal_b6

                    autofit_worksheet_columns(ws6_prod)
                    
                    output_b6_prod = io.BytesIO()
                    wb6_prod.save(output_b6_prod)
                    output_b6_prod.seek(0)
                    productivity_files["B6"] = output_b6_prod.getvalue()
                
                combined_output = io.BytesIO()
                with pd.ExcelWriter(combined_output, engine='openpyxl') as writer:
                    for bucket_name, filtered_df in processed_dfs.items():
                        output_df = filtered_df.drop(columns=["Card Number"])
                        output_df.to_excel(writer, index=False, sheet_name=bucket_name)
                combined_output.seek(0)
                
                temp_filename = f"temp_daily_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                b5_filename = f"AGENCY DAILY REPORT B5 AS OF {current_date}.xlsx"
                b6_filename = f"AGENCY DAILY REPORT B6 AS OF {current_date}.xlsx"
                b5_prod_filename = f"B5 Daily Productivity AS OF {current_date}.xlsx"
                b6_prod_filename = f"B6 Daily Productivity AS OF {current_date}.xlsx"
                
                return {
                    "b5_df": bucket5_df,
                    "b6_df": bucket6_df,
                    "b5_prod_df": b5_prod_df,
                    "b6_prod_df": b6_prod_df,
                    "b5_binary": b5_binary.getvalue() if not bucket5_df.empty else None,
                    "b6_binary": b6_binary.getvalue() if not bucket6_df.empty else None,
                    "b5_filename": b5_filename,
                    "b6_filename": b6_filename,
                    "b5_prod_binary": productivity_files.get("B5"),
                    "b6_prod_binary": productivity_files.get("B6"),
                    "b5_prod_filename": b5_prod_filename,
                    "b6_prod_filename": b6_prod_filename,
                    "preview": combined_output.getvalue(),
                    "temp_filename": temp_filename,
                    "output_files": output_files,
                    "productivity_files": productivity_files,
                    "output_filenames": {
                        "B5": b5_filename,
                        "B6": b6_filename,
                        "B5_Productivity": b5_prod_filename,
                        "B6_Productivity": b6_prod_filename
                    }
                }
                            
            return None, None, None
            
        except Exception as e:
            st.error(f"Error processing agency daily report: {str(e)}")
            return None, None, None

class SumishoProcessor(BaseProcessor):
    def process_daily_remark(self, file_content, sheet_name=None, preview_only=False,
    remove_duplicates=False, remove_blanks=False, trim_spaces=False,
    template_content=None, template_sheet=None, target_column=None):

        try:
            byte_stream = io.BytesIO(file_content)
            xls = pd.ExcelFile(byte_stream)
            df = pd.read_excel(xls, sheet_name=sheet_name)
            df = self.clean_data(df, remove_duplicates, remove_blanks, trim_spaces)

            if 'Date' not in df.columns or 'Remark' not in df.columns or 'Account No.' not in df.columns:
                raise ValueError("Required columns not found in the uploaded file.")
            
            df = df[df['Account No.'].notna()]                          
            df = df[df['Account No.'].astype(str).str.strip() != '']  

            if 'Time' in df.columns:
                if pd.api.types.is_object_dtype(df['Time']):
                    try:
                        df['Time'] = pd.to_datetime(df['Time'], format='%I:%M:%S %p')
                    except ValueError:
                        pass
                df = df.sort_values(by='Time', ascending=False)
                df = df.drop_duplicates(subset='Account No.', keep='first')
                
            df['FormattedDate'] = pd.to_datetime(df['Date']).dt.strftime('%m/%d/%Y')
            df['Date_Remark'] = df['FormattedDate'] + ' ' + df['Remark'].astype(str)
            
            account_remark_map = {}
            for idx, row in df.iterrows():
                account_number = str(int(row['Account No.']))
                formatted_date = row.get('FormattedDate')
                remark = row.get('Remark', '')

                if pd.isna(formatted_date):
                    value = str(remark) if pd.notna(remark) else ""
                else:
                    value = str(formatted_date) + ' ' + (str(remark) if pd.notna(remark) else "")
                    
                account_remark_map[account_number.strip()] = value
            
            if preview_only:
                template_stream = io.BytesIO(template_content)
                template_xls = pd.ExcelFile(template_stream)
                template_df = pd.read_excel(template_xls, sheet_name=template_sheet, header=1)
                
                account_number_col = None
                for col in template_df.columns:
                    col_str = str(col)
                    if 'ACCOUNT' in col_str.upper() and ('NUMBER' in col_str.upper() or 'NO' in col_str.upper()):
                        account_number_col = col
                        break
                        
                if not account_number_col:
                    st.write("Available columns:", template_df.columns.tolist())
                    raise ValueError("Account number column not found in template file.")
                    
                updated_count = 0
                for idx, row in template_df.iterrows():
                    template_acct = str(row[account_number_col]).strip() if pd.notna(row[account_number_col]) else ""
                    if template_acct in account_remark_map:
                        template_df.loc[idx, target_column] = account_remark_map[template_acct]
                        updated_count += 1
                
                st.write(f"Preview: {updated_count} cells would be updated in the template")
                return template_df
            
            output_filename = "Processed_Daily_Remark.xlsx"
            output_path = os.path.join(self.temp_dir, output_filename)
            
            template_stream = io.BytesIO(template_content)
            workbook = load_workbook(template_stream)
            
            if template_sheet in workbook.sheetnames:
                sheet = workbook[template_sheet]
            else:
                sheet = workbook.active
                st.warning(f"Sheet '{template_sheet}' not found, using active sheet instead")
            
            header_row = 2 
            account_col_idx = None
            target_col_idx = None
            
            for col_idx, cell in enumerate(sheet[header_row], 1):
                cell_value = str(cell.value).upper() if cell.value else ""
                if cell_value and ('ACCOUNT' in cell_value and ('NUMBER' in cell_value or 'NO' in cell_value)):
                    account_col_idx = col_idx
                if cell.value == target_column:
                    target_col_idx = col_idx

            if account_col_idx is None or target_col_idx is None:
                st.write("Header row content:", [cell.value for cell in sheet[header_row]])
                st.write(f"Looking for account column and target column: '{target_column}'")
                raise ValueError("Could not locate columns in Excel sheet")
                
            update_count = 0
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                account_cell = sheet.cell(row=row_idx, column=account_col_idx)
                
                if account_cell.value is not None:
                    account_str = str(account_cell.value).strip()
                    
                    if account_str in account_remark_map:
                        sheet.cell(row=row_idx, column=target_col_idx).value = account_remark_map[account_str]
                        update_count += 1

            st.write(f"Updated {update_count} cells in the Excel file")
            workbook.save(output_path)

            with open(output_path, 'rb') as f:
                output_binary = f.read()

            return None, output_binary, output_filename

        except Exception as e:
            st.error(f"Error processing daily report: {str(e)}")
            import traceback
            st.write(traceback.format_exc())
            raise

CAMPAIGN_CONFIG = {
    "No Campaign": {
        "automation_options": ["Data Clean"],
        "automation_map": {
            "Data Clean": "clean_only",
        },
        "processor": BaseProcessor    
    },
    "BPI": {
        "automation_options": ["Updates", "Uploads", "Cured List"],
        "automation_map": {
            "Uploads": "process_uploads",
            "Updates": "process_updates",
            "Cured List": "process_cured_list"
        },
        "processor": BPIProcessor
    },
    "ROB Bike": {
        "automation_options": ["Daily Remark Report", "Endorsement"],
        "automation_map": {
            "Daily Remark Report": "process_daily_remark",
            "Endorsement": "process_new_endorsement", 
        },
        "processor": ROBBikeProcessor
    },
    "BDO Auto B5 & B6": {
        "automation_options": ["Agency Daily Report", "Endorsement"],
        "automation_map": {
            "Agency Daily Report": "process_agency_daily_report",
            "Endorsement": "process_new_endorsement", 
        },
        "processor": BDOAutoProcessor
    },
    "Sumisho": {
        "automation_options": ["Daily Remark Report"],
        "automation_map": {
            "Daily Remark Report": "process_daily_remark",
        },
        "processor": SumishoProcessor
    },
}

def main():
    st.set_page_config(
        page_title="Automation Tool",
        layout="wide")
    
    st.markdown("""
        <style>
            .title {
                font-size: 24px;
                font-weight: bold;
            }
            .sub-title {
                font-size: 12px;
                margin-bottom: 15px;
            }
            div[data-baseweb] {
                font-size: 12px;
                line-height: 1.6 !important;
            }
            div[data-testid="stToolbar"] {
                display: none;
            }
            div[data-testid="stFileUploaderDropzoneInstructions"] {
                display: none;
            }
            section[data-testid="stFileUploaderDropzone"] {
                padding: 0px;
                margin: 0px;
                font-size: 12px;
            }
            button[data-testid="stBaseButton-secondary"] {
                width: 100%;
            }
        </style>
    """, unsafe_allow_html=True)
    
    st.markdown("<div class='title'>Automation Tool</div>", unsafe_allow_html=True)
    st.markdown("<div class='sub-title'>Transform Files into CMS Format</div>", unsafe_allow_html=True)

    campaign = st.sidebar.selectbox("Select Campaign", CAMPAIGN_CONFIG, index=0)
    config = CAMPAIGN_CONFIG[campaign]
    processor = config["processor"]()
    automation_map = config["automation_map"]
    automation_options = config["automation_options"]

    st.sidebar.header("Settings")
    automation_type = st.sidebar.selectbox("Select Automation Type", automation_options, key=f"{campaign}_automation_type")

    preview = st.sidebar.checkbox("Preview file before processing", value=True, key=f"{campaign}_preview")
    st.markdown("""

        """, unsafe_allow_html=True)
    uploaded_file = st.sidebar.file_uploader(
        "Upload File", 
        type=["xlsx", "xls"], 
        key=f"{campaign}_file_uploader"
    )
    
    if campaign == "ROB Bike" and automation_type == "Daily Remark Report":
        yesterday = datetime.now() - timedelta(days=1)
        report_date = st.sidebar.date_input('Date Report', value=yesterday, format="MM/DD/YYYY") 
        
        with st.sidebar.expander("Upload Other File", expanded=False):
            upload_field_result = st.file_uploader(
                "Field Result",
                type=["xlsx", "xls"],
                key=f"{campaign}_field_result"
            )
            upload_dataset = st.file_uploader(
                "Dataset",
                type=["xlsx", "xls"],
                key=f"{campaign}_dataset"
            )
            upload_disposition = st.file_uploader(
                "Disposition",
                type=["xlsx", "xls"],
                key=f"{campaign}_disposition"
            )
            
        if upload_field_result:
            TABLE_NAME = 'rob_bike_field_result'
            
            try:
                xls = pd.ExcelFile(upload_field_result)

                sheet_options = xls.sheet_names
                if len(sheet_options) > 1: 
                    selected_sheet = st.selectbox(
                        "Select a sheet from the Excel file:",
                        options=sheet_options,
                        index=0,
                        key="field_result_sheet_select"
                    )
                else:
                    selected_sheet = sheet_options[0]
                    
                if selected_sheet:
                    df = pd.read_excel(xls, sheet_name=selected_sheet)
                    df_clean = df.replace({np.nan: 0})
                
                if 'chcode' in df_clean.columns and 'status' in df_clean.columns and 'SUB STATUS' in df_clean.columns and 'DATE' in df_clean.columns and 'TIME' in df_clean.columns:
                    df_filtered = df_clean[(df_clean['status'] != 'CANCEL') & (df_clean['bank'] == 'ROB MOTOR LOAN')]
                    df_extracted = df_filtered[['chcode', 'status', 'SUB STATUS', 'DATE', 'TIME']].copy()
                    
                    df_extracted = df_extracted.rename(columns={
                        'SUB STATUS': 'substatus',
                        'DATE': 'date',
                        'TIME': 'time'
                    })
                    
                    df_extracted.loc[:, 'time'] = df_extracted['time'].astype(str).replace('NaT', '')
            
                    try:
                        temp_dates = pd.to_datetime(df_extracted['date'], errors='coerce')
                        df_extracted.loc[:, 'date'] = temp_dates.astype(str).str.split(' ').str[0]
                        df_extracted.loc[:, 'date'] = df_extracted['date'].replace('NaT', '')
                    except:
                        df_extracted.loc[:, 'date'] = df_extracted['date'].astype(str).replace('NaT', '')

                    df_extracted['inserted_date'] = pd.to_datetime(
                        df_extracted['date'].astype(str) + ' ' + df_extracted['time'].astype(str), 
                        errors='coerce'
                    )

                    df_extracted['inserted_date'] = df_extracted['inserted_date'].astype(str).replace('NaT', None)
                    
                    st.subheader("Extracted Records:")
                    st.dataframe(df_extracted)
                    
                    button_placeholder = st.empty()
                    upload_button = button_placeholder.button("Upload Records to Database", key="upload_button")
                    
                    if upload_button:
                        with st.spinner("Checking for existing records in database..."):
                            df_to_check = df_extracted.copy()
                            
                            unique_combinations = df_to_check[['chcode', 'status', 'date', 'time', 'inserted_date']].drop_duplicates()
                            
                            existing_records = []
                            total_combinations = len(unique_combinations)
                            
                            if total_combinations > 0:
                                check_progress = st.progress(0)
                                check_status = st.empty()
                                check_status.text(f"Checking 0 of {total_combinations} records...")
                                
                                batch_size = 100
                                for i in range(0, total_combinations, batch_size):
                                    batch = unique_combinations.iloc[i:i+batch_size]
                                    
                                    for _, row in batch.iterrows():
                                        chcode = row['chcode']
                                        status = row['status']
                                        inserted_date = row['inserted_date']
                                        
                                        query = supabase.table(TABLE_NAME).select("*").eq('chcode', chcode).eq('status', status)
                                        
                                        if inserted_date is not None and inserted_date != 'NaT':
                                            query = query.eq('inserted_date', inserted_date)
                                            
                                        try:
                                            response = query.execute()
                                            if hasattr(response, 'data') and response.data:
                                                existing_records.extend(response.data)
                                        except Exception as e:
                                            st.warning(f"Error checking record: {str(e)}. Continuing...")
                                    
                                    progress_value = min(1.0, (i + batch_size) / total_combinations)
                                    check_progress.progress(progress_value)
                                    check_status.text(f"Checking {min(i + batch_size, total_combinations)} of {total_combinations} records...")
                                
                                check_progress.empty()
                                check_status.empty()
                            
                            existing_df = pd.DataFrame(existing_records) if existing_records else pd.DataFrame()
                            
                            if not existing_df.empty:
                                df_extracted['chcode'] = df_extracted['chcode'].astype(str)
                                df_extracted['status'] = df_extracted['status'].astype(str)
                                
                                existing_df['chcode'] = existing_df['chcode'].astype(str)
                                existing_df['status'] = existing_df['status'].astype(str)
                                
                                df_extracted['unique_key'] = df_extracted['chcode'] + '_' + df_extracted['status'] + '_' + df_extracted['inserted_date'].astype(str)
                                
                                existing_keys = []
                                for _, row in existing_df.iterrows():
                                    key = str(row['chcode']) + '_' + str(row['status']) + '_' + str(row['inserted_date'])
                                    existing_keys.append(key)
                                
                                df_new_records = df_extracted[~df_extracted['unique_key'].isin(existing_keys)].copy()
                                df_new_records.drop('unique_key', axis=1, inplace=True)
                            else:
                                df_new_records = df_extracted.copy()
                        
                        total_records = len(df_extracted)
                        new_records = len(df_new_records)
                        duplicate_records = total_records - new_records
                        
                        st.info(f"Found {total_records} total records. {new_records} are new and {duplicate_records} already exist.")
                        
                        if new_records > 0:
                            try:
                                df_to_upload = df_new_records.copy()
                                
                                for col in df_to_upload.columns:
                                    if pd.api.types.is_datetime64_any_dtype(df_to_upload[col]):
                                        df_to_upload[col] = df_to_upload[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                                
                                df_to_upload = df_to_upload.astype(object).where(pd.notnull(df_to_upload), None)
                                
                                records_to_insert = df_to_upload.to_dict(orient="records")
                                
                                if records_to_insert:
                                    batch_size = 100
                                    success_count = 0
                                    
                                    progress_bar = st.progress(0)
                                    status_text = st.empty()
                                    
                                    for i in range(0, len(records_to_insert), batch_size):
                                        batch = records_to_insert[i:i+batch_size]
                                        
                                        if batch:
                                            try:
                                                response = supabase.table(TABLE_NAME).insert(batch).execute()
                                                
                                                if hasattr(response, 'data') and response.data:
                                                    success_count += len(response.data)
                                            except Exception as e:
                                                st.error(f"Error inserting batch: {str(e)}")
                                        
                                        progress = min(i + batch_size, len(records_to_insert)) / max(1, len(records_to_insert))
                                        progress_bar.progress(progress)
                                        status_text.text(f"Uploaded {success_count} of {len(records_to_insert)} records...")
                                    
                                    st.toast(f"Field Result Updated! {success_count} unique records uploaded successfully.")
                                    st.success("Upload completed successfully!")
                                else:
                                    st.warning("No new records to upload.")
                            
                            except Exception as e:
                                st.error(f"Error uploading field result: {str(e)}")
                                import traceback
                                st.code(traceback.format_exc())
                        else:
                            st.warning("No new records to upload. All records already exist in the database.")

                else:
                    st.error("Required columns not found in the uploaded file.")
            except Exception as e:
                st.error(f"Error processing Excel file: {str(e)}")
                
        if upload_dataset:
            TABLE_NAME = 'rob_bike_dataset'
            try:
                xls = pd.ExcelFile(upload_dataset)
                
                sheet_options = xls.sheet_names
                if len(sheet_options) > 1:
                    selected_sheet = st.selectbox(
                        "Select a sheet from the Excel file:",
                        options=sheet_options,
                        index=0,
                        key="dataset_sheet_select"
                    )
                else:
                    selected_sheet = sheet_options[0]
                    
                if selected_sheet:     
                    df = pd.read_excel(xls, sheet_name=selected_sheet)
                    df_clean = df.replace({np.nan: 0})
                    df_filtered = df_clean.copy()
                
                st.subheader("Uploaded Dataset:")
                st.dataframe(df_filtered)
                
                possible_column_variants = {
                    'ChCode': ['ChCode'],
                    'Account Number': ['Account Number', 'Account_Number'],
                    'Client Name': ['Client Name', 'Client_Name'],
                    'Endorsement Date': ['Endorsement Date', 'Endorsement_Date'],
                    'Endrosement DPD': ['Endrosement DPD', 'Endrosement_DPD'],
                    'Store': ['Store'],
                    'Cluster': ['Cluster']
                }
                
                target_columns = [
                    'chcode',
                    'account_number',
                    'client_name',
                    'endo_date',
                    'endo_dpd',
                    'stores',
                    'cluster'
                ]
                
                column_mapping = {}
                for (key, variants), target in zip(possible_column_variants.items(), target_columns):
                    for variant in variants:
                        if variant in df_filtered.columns:
                            column_mapping[variant] = target
                            break 
                        
                if len(column_mapping) == len(target_columns):
                    df_selected = df_filtered[list(column_mapping.keys())].rename(columns=column_mapping)
                    
                    df_selected = df_selected.rename(columns=column_mapping)
                    
                    button_placeholder = st.empty()
                    status_placeholder = st.empty()
                    
                    upload_button = button_placeholder.button("Upload to Database", key="upload_dataset_button")
                    
                    if upload_button:
                        button_placeholder.button("Processing...", disabled=True, key="processing_dataset_button")
                        
                        try:
                            unique_id_col = 'account_number'
                            unique_ids = df_selected[unique_id_col].astype(str).str.strip().unique().tolist()
                            
                            for col in df_selected.columns:
                                if pd.api.types.is_datetime64_any_dtype(df_selected[col]):
                                    df_selected[col] = df_selected[col].dt.strftime('%Y-%m-%d')
                            
                            df_selected = df_selected.astype(object).where(pd.notnull(df_selected), None)
                            df_selected[unique_id_col] = df_selected[unique_id_col].astype(str).str.strip() 
                            
                            new_records = df_selected.to_dict(orient="records")
                            
                            existing_records = []
                            batch_size_for_query = 20
                            
                            progress_bar = st.progress(0)
                            status_text = status_placeholder.empty()
                            status_text.text("Fetching existing records...")
                            
                            for i in range(0, len(unique_ids), batch_size_for_query):
                                batch_ids = unique_ids[i:i+batch_size_for_query]
                                batch_ids = [id for id in batch_ids if id is not None and id != '']
                                
                                if batch_ids:
                                    try:
                                        batch_response = supabase.table(TABLE_NAME).select("*").in_(unique_id_col, batch_ids).execute()
                                        if hasattr(batch_response, 'data') and batch_response.data:
                                            existing_records.extend(batch_response.data)
                                    except Exception as e:
                                        st.warning(f"Error fetching batch {i}: {str(e)}. Continuing...")
                                
                                progress_value = min(1.0, (i + batch_size_for_query) / max(1, len(unique_ids)))
                                progress_bar.progress(progress_value)
                            
                            existing_df = pd.DataFrame(existing_records) if existing_records else pd.DataFrame()
                            if not existing_df.empty:
                                existing_df[unique_id_col] = existing_df[unique_id_col].astype(str).str.strip()
                            
                            records_to_insert = []
                            records_to_update = []
                            total_records = len(new_records)
                            processed_count = 0
                            
                            status_text.text("Identifying records to insert or update...")
                            progress_bar.progress(0)
                            
                            def records_differ(new_record, existing_record):
                                for key, value in new_record.items():
                                    if key in existing_record and str(value).strip() != str(existing_record[key]).strip():
                                        return True
                                return False
                            
                            for new_record in new_records:
                                processed_count += 1
                                account_number = str(new_record[unique_id_col]).strip()
                                
                                if not existing_df.empty:
                                    matching_records = existing_df[existing_df[unique_id_col] == account_number]
                                    
                                    if not matching_records.empty:
                                        existing_record = matching_records.iloc[0].to_dict()
                                        if records_differ(new_record, existing_record):
                                            new_record['id'] = existing_record['id']
                                            records_to_update.append(new_record)
                                    else:
                                        records_to_insert.append(new_record)
                                else:
                                    records_to_insert.append(new_record)
                                
                                progress_value = min(1.0, processed_count / total_records)
                                progress_bar.progress(progress_value)
                            
                            status_placeholder.info(f"Found {len(records_to_insert)} records to insert and {len(records_to_update)} records to update.")
                            
                            batch_size_for_db = 100
                            success_count = 0
                            
                            if records_to_insert:
                                status_text.text("Inserting new records...")
                                progress_bar.progress(0)
                                
                                for i in range(0, len(records_to_insert), batch_size_for_db):
                                    batch = records_to_insert[i:i+batch_size_for_db]
                                    
                                    if batch:
                                        try:
                                            response = supabase.table(TABLE_NAME).insert(batch).execute()
                                            if hasattr(response, 'data') and response.data:
                                                success_count += len(batch)
                                        except Exception as e:
                                            st.error(f"Error inserting records batch: {str(e)}")
                                    
                                    progress_value = min(1.0, min(i + batch_size_for_db, len(records_to_insert)) / max(1, len(records_to_insert)))
                                    progress_bar.progress(progress_value)
                                    status_text.text(f"Inserted {success_count} of {len(records_to_insert)} new records...")
                            
                            update_count = 0
                            if records_to_update:
                                status_text.text("Updating existing records...")
                                progress_bar.progress(0)
                                
                                for i, record in enumerate(records_to_update):
                                    record_id = record.pop('id')
                                    
                                    try:
                                        response = supabase.table(TABLE_NAME).update(record).eq('id', record_id).execute()
                                        if hasattr(response, 'data') and response.data:
                                            update_count += 1
                                    except Exception as e:
                                        st.error(f"Error updating record {record_id}: {str(e)}")
                                    
                                    progress_value = min(1.0, (i + 1) / len(records_to_update))
                                    progress_bar.progress(progress_value)
                                    status_text.text(f"Updated {update_count} of {len(records_to_update)} existing records...")
                            
                            total_processed = success_count + update_count
                            if total_processed > 0:
                                st.toast(f"Dataset Updated! {success_count} records inserted and {update_count} records updated successfully.")
                                button_placeholder.button("Upload Complete!", disabled=True, key="complete_dataset_button")
                            else:
                                st.warning("No records were processed. Either no changes were needed or the operation failed.")
                                button_placeholder.button("Try Again", key="retry_dataset_button")
                        
                        except Exception as e:
                            st.error(f"Error uploading dataset: {str(e)}")
                            import traceback
                            st.code(traceback.format_exc())
                            button_placeholder.button("Upload Failed - Try Again", key="error_dataset_button")
                else:
                    missing_cols = [col for col in possible_column_variants if col not in df_filtered.columns]
                    st.error(f"Required columns not found in the uploaded file.")
                    
            except Exception as e:
                st.error(f"Error processing Excel file: {str(e)}")
                        
        if upload_disposition:
            TABLE_NAME = 'rob_bike_disposition'
            try:
                xls = pd.ExcelFile(upload_disposition)
                
                sheet_options = xls.sheet_names
                if len(sheet_options) > 1:
                    selected_sheet = st.selectbox(
                        "Select a sheet from the Excel file:",
                        options=sheet_options,
                        index=0,
                        key="disposition_sheet_select"
                    )
                else:
                    selected_sheet = sheet_options[0]    
                    
                if selected_sheet:
                    df = pd.read_excel(xls, sheet_name=selected_sheet)
                    df_clean = df.replace({np.nan: ''})
                    df_filtered = df_clean.copy()

                st.subheader("Uploaded Disposition:")
                st.dataframe(df_filtered)

                button_placeholder = st.empty()
                upload_button = button_placeholder.button("Upload to Database", key="upload_disposition_button")

                if upload_button:
                    button_placeholder.button("Processing...", disabled=True, key="processing_disposition_button")
                    try:
                        if 'CMS Disposition' in df_filtered.columns:
                            unique_dispositions = df_filtered['CMS Disposition'].drop_duplicates().tolist()

                            existing_response = supabase.table(TABLE_NAME).select("disposition").execute()
                            if existing_response.data is None:
                                existing_dispositions = []
                            else:
                                existing_dispositions = [record['disposition'] for record in existing_response.data]

                            records_to_insert = [
                                {"disposition": d} for d in unique_dispositions if d not in existing_dispositions
                            ]

                            if records_to_insert:
                                insert_response = supabase.table(TABLE_NAME).insert(records_to_insert).execute()
                                toast_placeholder = st.empty()
                                toast_placeholder.success("Upload successful!")
                                time.sleep(3)
                                toast_placeholder.empty()
                            else:
                                st.info("No new dispositions to add; all values already exists.")

                            button_placeholder.empty()
                        else:
                            st.error("Required columns was not found in the uploaded file.")
                    except Exception as e:
                        st.error(f"Error uploading disposition: {str(e)}")
                        button_placeholder.button("Upload Failed - Try Again", key="error_disposition_button")        
            except Exception as e:
                st.error(f"An error occurred: {str(e)}")

    if campaign == "BDO Auto B5 & B6" and automation_type == "Agency Daily Report":
        def clean_number_input(label):
            raw_input = st.sidebar.text_input(label)
            clean_input = raw_input.replace(",", "")
            try:
                return float(clean_input)
            except ValueError:
                return None
                
        st.sidebar.subheader("B5")
        col1, col2 = st.columns(2)
        with col1:
            kept_count_b5 = clean_number_input("Kept Count (B5)")
        with col2:
            kept_bal_b5 = clean_number_input("Kept Balance (B5)")
        alloc_bal_b5 = clean_number_input("Allocation Balance (B5)")

        st.sidebar.subheader("B6")
        col1, col2 = st.columns(2)
        with col1:
            kept_count_b6 = clean_number_input("Kept Count (B6)")
        with col2:
            kept_bal_b6 = clean_number_input("Kept Balance (B6)")
        alloc_bal_b6 = clean_number_input("Allocation Balance (B6)")

    if campaign == "Sumisho" and automation_type == "Daily Remark Report":
        upload_madrid_daily = st.sidebar.file_uploader(
            "SP Madrid Daily",
            type=["xlsx", "xls"],
            key=f"{campaign}_sp_madrid_daily"
        )

        if upload_madrid_daily is not None:
            sp_madrid_daily = upload_madrid_daily.getvalue()

            template_stream = io.BytesIO(sp_madrid_daily)
            template_xls = pd.ExcelFile(template_stream)
            template_sheets = template_xls.sheet_names

            selected_template_sheet = st.sidebar.selectbox("Select a sheet from the SP Madrid Daily Template", template_sheets)

            template_stream.seek(0)
            template_df_preview = pd.read_excel(template_stream, sheet_name=selected_template_sheet, header=1)
            available_columns = list(template_df_preview.columns)

            selected_date_column = st.sidebar.selectbox("Select the column to insert 'Date + Remark'", available_columns)

        else:
            st.warning("Please upload the SP Madrid Daily template file.")
            st.stop()
            

    df = None
    sheet_names = []

    if uploaded_file is not None:
        if 'previous_filename' not in st.session_state or st.session_state['previous_filename'] != uploaded_file.name:
            if 'output_binary' in st.session_state:
                del st.session_state['output_binary']
            if 'output_filename' in st.session_state:
                del st.session_state['output_filename']
            if 'result_sheet_names' in st.session_state:
                del st.session_state['result_sheet_names']
                
            st.session_state['previous_filename'] = uploaded_file.name
        
        with st.sidebar.expander("Data Cleaning Options"):
            remove_duplicates = st.checkbox("Remove Duplicates", value=False, key=f"{campaign}_remove_duplicates")
            remove_blanks = st.checkbox("Remove Blanks", value=False, key=f"{campaign}_remove_blanks")
            trim_spaces = st.checkbox("Trim Text", value=False, key=f"{campaign}_trim_spaces")
        
        with st.sidebar.expander("Data Manipulation"):
            st.markdown("#### Column Operations")
            enable_add_column = st.checkbox("Add Column", value=False)
            enable_column_removal = st.checkbox("Remove Column", value=False)
            enable_column_renaming = st.checkbox("Rename Column", value=False)
            
            st.markdown("#### Row Operations")
            enable_row_filtering = st.checkbox("Filter Row", value=False)
            enable_add_row = st.checkbox("Add Row", value=False)
            enable_row_removal = st.checkbox("Remove Row", value=False)
            
            st.markdown("#### Value Operations")
            enable_edit_values = st.checkbox("Edit Values", value=False)
          
        file_content = uploaded_file.getvalue()
        file_buffer = io.BytesIO(file_content)
        
        try:
            xlsx = pd.ExcelFile(file_buffer)
            sheet_names = xlsx.sheet_names
            is_encrypted = False
            decrypted_file = file_buffer
            
        except Exception as e:
            if "file has been corrupted" in str(e) or "Workbook is encrypted" in str(e):
                is_encrypted = True
                st.sidebar.warning("This file appears to be password protected.")
                excel_password = st.sidebar.text_input("Enter Excel password", type="password")
                
                if not excel_password:
                    st.warning("Please enter the Excel file password.")
                    st.stop()
                
                try:
                    decrypted_file = io.BytesIO()
                    office_file = msoffcrypto.OfficeFile(io.BytesIO(file_content))
                    office_file.load_key(password=excel_password)
                    office_file.decrypt(decrypted_file)
                    decrypted_file.seek(0)
                    xlsx = pd.ExcelFile(decrypted_file)
                    sheet_names = xlsx.sheet_names
                except Exception as decrypt_error:
                    st.sidebar.error(f"Decryption failed: {str(decrypt_error)}")
                    st.stop()
            else:
                st.sidebar.error(f"Error reading file: {str(e)}")
                st.stop()
        
        if len(sheet_names) > 1 :
            selected_sheet = st.sidebar.selectbox(
                "Select Sheet", 
                options=sheet_names,
                index=0,
                key=f"{campaign}_sheet_selector"
            )
        else:
            selected_sheet = sheet_names[0]
        
        try:
            if is_encrypted:
                decrypted_file.seek(0)
                df = pd.read_excel(decrypted_file, sheet_name=selected_sheet)
            else:
                df = pd.read_excel(xlsx, sheet_name=selected_sheet)
                
            if selected_sheet and preview:
                st.subheader(f"Preview of {selected_sheet}")
                df_preview = df.copy().dropna(how='all').dropna(how='all', axis=1)
                st.dataframe(df_preview, use_container_width=True)
                
        except Exception as e:
            st.sidebar.error(f"Error reading sheet: {str(e)}")
    
    process_button = st.sidebar.button("Process File", type="primary", disabled=uploaded_file is None, key=f"{campaign}_process_button")

    if uploaded_file is not None:
        file_content = uploaded_file.getvalue() if hasattr(uploaded_file, 'getvalue') else uploaded_file.read()
        
        try:
            if "renamed_df" in st.session_state:
                df = st.session_state["renamed_df"]
            else:
                pass
            
            df = df.dropna(how='all', axis=0) 
            df = df.dropna(how='all', axis=1)

            if enable_add_column:
                st.subheader("Add New Columns")

                if "column_definitions" not in st.session_state:
                    st.session_state.column_definitions = []

                with st.form("add_column_form", clear_on_submit=True):
                    new_column_name = st.text_input("New Column Name")
                    column_source_type = st.radio("Column Source", ["Input Value", "Copy From Column", "Excel-like Formula"], key="source_type")

                    source_column = modification_type = prefix_text = suffix_text = selected_function = custom_function = formula = None
                    
                    if column_source_type == "Input Value":
                        input_value = st.text_input("Value to fill in each row")
                    elif column_source_type == "Copy From Column":
                        source_column = st.selectbox("Source Column (copy from)", df.columns.tolist(), key="source_column")
                        modification_type = st.radio("Modification Type", ["Direct Copy", "Text Prefix", "Text Suffix", "Apply Function"], key="mod_type")

                        if modification_type == "Text Prefix":
                            prefix_text = st.text_input("Prefix to add")
                        elif modification_type == "Text Suffix":
                            suffix_text = st.text_input("Suffix to add")
                        elif modification_type == "Apply Function":
                            function_options = ["To Uppercase", "To Lowercase", "Strip Spaces", "Custom Function"]
                            selected_function = st.selectbox("Select Function", function_options)
                            if selected_function == "Custom Function":
                                custom_function = st.text_area("Custom function (use 'x')", value="lambda x: x")
                    else:
                        st.info("Use column names in curly braces {} and expressions (e.g. `{Amount} * 2`, etc.)")
                        formula = st.text_area("Excel-like formula", height=80)

                    submitted = st.form_submit_button("Add to List")
                    if submitted and new_column_name:
                        st.session_state.column_definitions.append({
                            "name": new_column_name,
                            "source": column_source_type,
                            "source_column": source_column,
                            "modification_type": modification_type,
                            "prefix_text": prefix_text,
                            "suffix_text": suffix_text,
                            "function": selected_function,
                            "custom_function": custom_function,
                            "formula": formula,
                            "input_value": input_value if column_source_type == "Input Value" else None,
                        })
                        st.success(f"Queued column: {new_column_name}")

                if st.session_state.column_definitions:
                    st.write("🧾 Queued Columns to Add:")
                    for idx, col_def in enumerate(st.session_state.column_definitions):
                        st.markdown(f"- **{col_def['name']}** from **{col_def['source']}**")

                    if st.button("Apply All Column Additions"):
                            
                        try:
                            for col_def in st.session_state.column_definitions:
                                name = col_def["name"]
                                source = col_def["source"]
                                
                                if source == "Input Value":
                                    input_value = col_def["input_value"]
                                    df[name] = input_value
                                elif source == "Copy From Column":
                                    source_col = col_def["source_column"]
                                    mod_type = col_def["modification_type"]

                                    if mod_type == "Direct Copy":
                                        df[name] = df[source_col]
                                    elif mod_type == "Text Prefix":
                                        df[name] = col_def["prefix_text"] + df[source_col].astype(str)
                                    elif mod_type == "Text Suffix":
                                        df[name] = df[source_col].astype(str) + col_def["suffix_text"]
                                    elif mod_type == "Apply Function":
                                        if col_def["function"] == "To Uppercase":
                                            df[name] = df[source_col].astype(str).str.upper()
                                        elif col_def["function"] == "To Lowercase":
                                            df[name] = df[source_col].astype(str).str.lower()
                                        elif col_def["function"] == "Strip Spaces":
                                            df[name] = df[source_col].astype(str).str.strip()
                                        elif col_def["function"] == "Custom Function":
                                            func = eval(col_def["custom_function"])
                                            df[name] = df[source_col].apply(func)

                                elif source == "Excel-like Formula":
                                    formula = col_def["formula"]
                                    processed = formula
                                    for col in df.columns:
                                        pattern = r'\{' + re.escape(col) + r'\}'
                                        processed = re.sub(pattern, f"df['{col}']", processed)
                                    processed = processed.replace("IF(", "np.where(").replace("SUM(", "np.sum(")
                                    processed = processed.replace("AVG(", "np.mean(").replace("MAX(", "np.max(").replace("MIN(", "np.min(")
                                    df[name] = eval(processed)

                            st.success("All queued columns added successfully!")
                            st.session_state.renamed_df = df
                            st.session_state.column_definitions.clear()
                        except Exception as e:
                            st.error(f"Error applying column additions: {str(e)}")

            if enable_column_removal:
                st.subheader("Column Removal")
                cols = df.columns.tolist()
                cols_to_remove = st.multiselect("Select columns to remove", cols)
                if cols_to_remove:
                    df = df.drop(columns=cols_to_remove)
                    st.success(f"Removed columns: {', '.join(cols_to_remove)}")

            if enable_column_renaming:
                st.subheader("Column Renaming")
                
                rename_df = pd.DataFrame({
                    "original_name": df.columns,
                    "new_name": df.columns
                })
                
                edited_df = st.data_editor(
                    rename_df,
                    column_config={
                        "original_name": st.column_config.TextColumn("Original Column Name", disabled=True),
                        "new_name": st.column_config.TextColumn("New Column Name")
                    },
                    hide_index=True,
                    key="column_rename_editor"
                )
                
                if st.button("Apply Column Renames", key="apply_multiple_renames"):
                    rename_dict = {
                        orig: new 
                        for orig, new in zip(edited_df["original_name"], edited_df["new_name"]) 
                        if orig != new
                    }

                    if rename_dict:
                        df = df.rename(columns=rename_dict)
                        st.session_state["renamed_df"] = df
                        st.success(f"Renamed {len(rename_dict)} column(s): {', '.join([f'{k} → {v}' for k, v in rename_dict.items()])}")

            if enable_row_filtering:
                st.subheader("Row Filtering")
                filter_col = st.selectbox("Select column to filter by", df.columns.tolist())
                filter_value = st.text_input("Enter search/filter value")
                
                if filter_value and filter_col:
                    if pd.api.types.is_numeric_dtype(df[filter_col]):
                        try:
                            filter_value_num = float(filter_value)
                            filtered_df = df[df[filter_col] == filter_value_num]
                        except ValueError:
                            st.warning("Entered value is not numeric. Using string comparison instead.")
                            filtered_df = df[df[filter_col].astype(str).str.contains(filter_value, case=False, na=False)]
                    else:
                        filtered_df = df[df[filter_col].astype(str).str.contains(filter_value, case=False, na=False)]

                    st.write(f"Found {len(filtered_df)} rows matching filter: '{filter_value}' in column '{filter_col}'")
                    df = filtered_df
                    
            if enable_add_row:
                st.subheader("Add New Rows")
                with st.form("add_row_form"):
                    row_data = {}
                    for col in df.columns:
                        row_data[col] = st.text_input(f"Value for {col}", "")
                    
                    add_row_submitted = st.form_submit_button("Add Row")
                    
                    if add_row_submitted:
                        new_row = pd.DataFrame([row_data])
                        df = pd.concat([df, new_row], ignore_index=True)
                        st.success("Row added successfully!")
                        st.session_state["renamed_df"] = df

            if enable_row_removal:
                st.subheader("Remove Rows")
                st.info("Select rows to remove by index")
                
                with st.form("remove_row_form"):
                    row_indices = st.multiselect("Select row indices to remove", 
                                                options=list(range(len(df))),
                                                format_func=lambda x: f"Row {x}")
                    
                    remove_rows_submitted = st.form_submit_button("Remove Selected Rows")
                    
                    if remove_rows_submitted and row_indices:
                        df = df.drop(index=row_indices).reset_index(drop=True)
                        st.success(f"Removed {len(row_indices)} row(s)")
                        st.session_state["renamed_df"] = df

            if enable_edit_values:
                st.subheader("Edit Values")
                
                edited_df = st.data_editor(
                    df,
                    num_rows="dynamic",
                    use_container_width=True,
                    key="value_editor"
                )
                
                if st.button("Apply Value Changes"):
                    st.session_state["renamed_df"] = edited_df
                    st.success("Value changes applied!")
                    
            if enable_add_column or enable_column_removal or enable_column_renaming or enable_row_filtering or enable_add_row or enable_row_removal or enable_edit_values:
                buffer = io.BytesIO()
                df.to_excel(buffer, index=False, engine='openpyxl')
                file_content = buffer.getvalue()
                st.subheader("Modified Data Preview")
                st.dataframe(df, use_container_width=True)

        except Exception as e:
            st.error(f"Error loading or manipulating file: {str(e)}")

        if "renamed_df" in st.session_state:
            df = st.session_state["renamed_df"]
            buffer = io.BytesIO()
            df.to_excel(buffer, index=False, engine='openpyxl')
            buffer.seek(0)
            file_content = buffer.getvalue()

        if process_button and selected_sheet:
            try:
                with st.spinner("Processing file..."):
                    if automation_type == "Cured List":
                        result = processor.process_cured_list(
                            file_content, 
                            sheet_name=selected_sheet,
                            preview_only=False,
                            remove_duplicates=remove_duplicates, 
                            remove_blanks=remove_blanks, 
                            trim_spaces=trim_spaces
                        )
                        st.session_state['cured_list_result'] = result
                        
                    elif automation_type == "Agency Daily Report":
                        if None in [kept_count_b5, kept_bal_b5, alloc_bal_b5, kept_count_b6, kept_bal_b6, alloc_bal_b6]:
                            st.error("Please enter valid numbers for all B5 and B6 fields (numbers only, commas allowed).")
                        else: 
                            result = processor.process_agency_daily_report(
                                file_content, 
                                sheet_name=selected_sheet,
                                preview_only=False,
                                remove_duplicates=remove_duplicates, 
                                remove_blanks=remove_blanks, 
                                trim_spaces=trim_spaces,
                                kept_count_b5=kept_count_b5,
                                kept_bal_b5=kept_bal_b5,
                                alloc_bal_b5=alloc_bal_b5,
                                kept_count_b6=kept_count_b6,
                                kept_bal_b6=kept_bal_b6,
                                alloc_bal_b6=alloc_bal_b6
                            )
                            st.session_state['agency_daily_result'] = result
                        
                        
                    else:
                        if automation_type == "Data Clean":
                            result_df, output_binary, output_filename = getattr(processor, automation_map[automation_type])(
                                file_content, 
                                sheet_name=selected_sheet,
                                preview_only=False,
                                remove_duplicates=remove_duplicates,
                                remove_blanks=remove_blanks,
                                trim_spaces=trim_spaces,
                                file_name=uploaded_file.name
                            )
                        elif campaign == "ROB Bike" and automation_type == "Daily Remark Report":
                            result_df, output_binary, output_filename = getattr(processor, automation_map[automation_type])(
                                file_content,  
                                sheet_name=selected_sheet,
                                preview_only=False,
                                remove_duplicates=remove_duplicates, 
                                remove_blanks=remove_blanks, 
                                trim_spaces=trim_spaces,
                                report_date=report_date
                            )
                        elif campaign == "Sumisho" and automation_type == "Daily Remark Report":
                            result_df, output_binary, output_filename = getattr(processor, automation_map[automation_type])(
                                file_content,  
                                sheet_name=selected_sheet,
                                preview_only=False,
                                remove_duplicates=remove_duplicates, 
                                remove_blanks=remove_blanks, 
                                trim_spaces=trim_spaces,
                                template_content=sp_madrid_daily,
                                template_sheet=selected_template_sheet,
                                target_column=selected_date_column
                            )
                        else:
                            result_df, output_binary, output_filename = getattr(processor, automation_map[automation_type])(
                                file_content, 
                                sheet_name=selected_sheet,
                                preview_only=False,
                                remove_duplicates=remove_duplicates,
                                remove_blanks=remove_blanks,
                                trim_spaces=trim_spaces
                            )
                            
                        if output_binary:
                            st.session_state['output_binary'] = output_binary
                            st.session_state['output_filename'] = output_filename
                            
                            excel_file = pd.ExcelFile(io.BytesIO(output_binary))
                            result_sheet_names = excel_file.sheet_names
                            st.session_state['result_sheet_names'] = result_sheet_names
                        
                        else:
                            st.error("No output file was generated")

                if "renamed_df" in st.session_state:
                    st.session_state.pop("renamed_df", None)

            except Exception as e:
                st.error(f"Error processing file: {str(e)}")

        if automation_type == "Cured List" and 'cured_list_result' in st.session_state:
            result = st.session_state['cured_list_result']
            if result != (None, None, None):
                tabs = st.tabs(["Remarks", "Reshuffle", "Payments"])
                with tabs[0]:
                    st.subheader("Remarks Data")
                    st.dataframe(result['remarks_df'], use_container_width=True)
                    st.download_button(label="Download Remarks File", data=result['remarks_binary'], file_name=result['remarks_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="remarks_download")
                with tabs[1]:
                    st.subheader("Reshuffle Data")
                    st.dataframe(result['others_df'], use_container_width=True)
                    st.download_button(label="Download Reshuffle File", data=result['others_binary'], file_name=result['others_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="reshuffle_download")
                with tabs[2]:
                    st.subheader("Payments Data")
                    st.dataframe(result['payments_df'], use_container_width=True)
                    st.download_button(label="Download Payments File", data=result['payments_binary'], file_name=result['payments_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="payments_download")
            else:
                st.error("Error processing. Please check the file content and try again.")
        elif automation_type == "Agency Daily Report" and 'agency_daily_result' in st.session_state:
            result = st.session_state['agency_daily_result']
            if result != (None, None, None):
                tabs = st.tabs(["Daily Report B5", "Daily Report B6", "B5 Prod", "B6 Prod"])
                with tabs[0]:
                    st.subheader("Daily Report B5")
                    st.dataframe(result['b5_df'], use_container_width=True)
                    st.download_button(label="Download Agency Daily Report B5 File", data=result['b5_binary'], file_name=result['b5_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="b5_download")
                with tabs[1]:
                    st.subheader("Daily Report B6")
                    st.dataframe(result['b6_df'], use_container_width=True)
                    st.download_button(label="Download Agency Daily Report B6 File", data=result['b6_binary'], file_name=result['b6_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="b6_download")
                with tabs[2]:
                    st.subheader("B5 Prod")
                    st.dataframe(result['b5_prod_df'], use_container_width=True)
                    st.download_button(label="Download Daily Productivity B5 Report File", data=result['b5_prod_binary'], file_name=result['b5_prod_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="b5_prod_download")
                with tabs[3]:
                    st.subheader("B6 Prod")
                    st.dataframe(result['b6_prod_df'], use_container_width=True)
                    st.download_button(label="Download Daily Productivity B6 Report File", data=result['b6_prod_binary'], file_name=result['b6_prod_filename'], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="b6_prod_download")
            else:
                st.error("Error processing. Please check the file content and try again.")
        elif 'output_binary' in st.session_state and 'result_sheet_names' in st.session_state:
            excel_file = pd.ExcelFile(io.BytesIO(st.session_state['output_binary']))
            result_sheet_names = st.session_state['result_sheet_names']
            
            if len(result_sheet_names) > 1:
                result_sheet = st.selectbox(
                    "Select Sheet",
                    options=result_sheet_names,
                    index=0,
                    key=f"{campaign}_result_sheet"
                )
            else: 
                result_sheet = result_sheet_names[0]
            
            selected_df = pd.read_excel(io.BytesIO(st.session_state['output_binary']), sheet_name=result_sheet)
            
            st.subheader("Processed Preview")
            st.dataframe(selected_df, use_container_width=True)
            
            st.download_button(
                label="Download File", 
                data=st.session_state['output_binary'], 
                file_name=st.session_state['output_filename'], 
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success(f"File processed successfully! Download '{st.session_state['output_filename']}'")

if __name__ == "__main__":
    main()